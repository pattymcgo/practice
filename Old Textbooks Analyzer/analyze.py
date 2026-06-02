"""
Old Textbooks Stacks Analyzer
==============================
Analyzes items in your library's Stacks locations for textbook likelihood
and flags items that should move to the Old Textbooks collection.

Can fetch data directly from Alma API, or read from an Alma Analytics CSV export.

Confidence levels:
  - OBVIOUS: Clear textbook signals (custom edition, 650 $v Textbooks, 655 Textbooks)
  - LIKELY:  Strong title/publisher/edition signals (score >= 50)
  - UNCLEAR: Weak or no signals -- needs human review

HOW TO RUN:

  From an Alma Analytics CSV export (recommended):
    python analyze.py --input data/Stacks_Oversized_Collection.csv
    python analyze.py --input data/Stacks_Oversized_Collection.csv --limit 100

  From the Alma API directly (requires Bibs search permission):
    python analyze.py              # Fetch fresh data from Alma + analyze
    python analyze.py --cached     # Skip API fetch, reanalyze from cached data
    python analyze.py --fetch-only # Only fetch and cache data, don't analyze yet

HOW TO EXPORT FROM ALMA ANALYTICS:
  1. In Alma, go to Analytics > Design Analytics
  2. Run a Physical Items report filtered to locations: STACK, OVERSIZE
  3. Include columns: MMS Id, Title, Author, Publisher, ISBN,
                      Publication Date, Location Name, Subjects
  4. Export as CSV and place in the data/ folder

OUTPUT:
    output/stacks_analysis_YYYYMMDD_HHMMSS.xlsx  -- timestamped spreadsheet
    output/latest_analysis.xlsx                   -- copy used by the web review app
    cache/production/stacks_raw.json              -- raw API data (speeds up re-runs)
    cache/production/stacks_enriched.json         -- enriched data with MARC + holdings
"""

import argparse
import json
import re
import sys
import time
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================================
# PATHS
# ============================================================================

BASE_DIR = Path(__file__).parent
CACHE_DIR = BASE_DIR / "cache"
OUTPUT_DIR = BASE_DIR / "output"

CACHE_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# ============================================================================
# CONFIGURATION
# ============================================================================

def load_config(sandbox=False):
    """Load settings from config.json (or config_sandbox.json) in the same folder."""
    filename    = "config_sandbox.json" if sandbox else "config.json"
    config_path = BASE_DIR / filename
    try:
        with open(config_path) as f:
            config = json.load(f)
        print(f"Config loaded: {config_path}")
        return config
    except FileNotFoundError:
        print(f"ERROR: {filename} not found at {config_path}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"ERROR: {filename} has invalid format: {e}")
        sys.exit(1)

# ============================================================================
# ALMA API
# ============================================================================

ALMA_BASE = "https://api-na.hosted.exlibrisgroup.com"


def make_headers(api_key):
    return {
        "Authorization": f"apikey {api_key}",
        "Accept": "application/json",
    }


def search_bibs_by_location(location_code, api_key, delay=0.3):
    """
    Fetch all bib records that have holdings in the given location code.
    Automatically pages through all results (50 at a time).

    Returns a list of brief bib dicts, each tagged with '_source_location'.
    """
    all_bibs = []
    offset = 0
    limit = 50
    total = None

    print(f"\n  Searching location: {location_code}")

    while True:
        params = {
            "q": f"holding_location~{location_code}",
            "limit": limit,
            "offset": offset,
            "view": "brief",
        }
        url = f"{ALMA_BASE}/almaws/v1/bibs"

        try:
            resp = requests.get(
                url, headers=make_headers(api_key), params=params, timeout=30
            )
        except requests.exceptions.RequestException as e:
            print(f"\n  Network error: {e}")
            break

        if resp.status_code == 200:
            data = resp.json()

            if total is None:
                total = int(data.get("total_record_count", 0))
                print(f"  Total records: {total:,}")
                if total == 0:
                    print(
                        f"  NOTE: 0 records found for location '{location_code}'.\n"
                        f"  Double-check that this location code exists in Alma."
                    )
                    break

            bibs = data.get("bib", [])
            if not isinstance(bibs, list):
                bibs = [bibs] if bibs else []

            for b in bibs:
                b["_source_location"] = location_code

            all_bibs.extend(bibs)
            print(f"  Fetched {len(all_bibs):,} / {total:,} ...", end="\r")

            if len(all_bibs) >= total or not bibs:
                break

            offset += limit
            time.sleep(delay)

        elif resp.status_code == 400:
            print(
                f"\n  Bad request (400) for location '{location_code}'.\n"
                f"  The Alma API may not support this query. Response:\n"
                f"  {resp.text[:400]}"
            )
            break
        else:
            print(f"\n  API error {resp.status_code}: {resp.text[:200]}")
            break

    print(f"\n  Done: {len(all_bibs):,} records from {location_code}")
    return all_bibs


def get_full_bib(mms_id, api_key):
    """
    Fetch the full bib record for an MMS ID.
    The full bib includes MARC data in the 'anies' field.
    """
    url = f"{ALMA_BASE}/almaws/v1/bibs/{mms_id}"
    try:
        resp = requests.get(url, headers=make_headers(api_key), timeout=30)
        if resp.status_code == 200:
            return resp.json()
    except requests.exceptions.RequestException:
        pass
    return None


def get_holdings(mms_id, api_key):
    """
    Fetch holdings records for a bib.
    Returns a list of holding dicts (each has location code and call number).
    """
    url = f"{ALMA_BASE}/almaws/v1/bibs/{mms_id}/holdings"
    try:
        resp = requests.get(url, headers=make_headers(api_key), timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            holdings = data.get("holding", [])
            return holdings if isinstance(holdings, list) else [holdings]
    except requests.exceptions.RequestException:
        pass
    return []

# ============================================================================
# MARC PARSING
# ============================================================================

def parse_marc_xml(marc_xml_str):
    """
    Parse a MARC XML string and pull out fields useful for textbook detection.

    Alma stores the MARC record as an XML string inside the 'anies' field
    of the bib API response. This function reads that XML.

    Returns a dict with:
        subjects             - list of subject heading strings (6xx fields)
        genre_forms          - list of genre/form strings (655)
        notes                - list of note strings (500, 504, 520, 521)
        has_textbook_subject - True if any 650 $v contains "Textbooks"
        has_textbook_genre   - True if any 655 contains "Textbook"
    """
    result = {
        "subjects": [],
        "genre_forms": [],
        "notes": [],
        "has_textbook_subject": False,
        "has_textbook_genre": False,
    }

    if not marc_xml_str or not isinstance(marc_xml_str, str):
        return result

    try:
        root = ET.fromstring(marc_xml_str)
    except ET.ParseError:
        return result

    # Walk every element -- handles both namespaced and plain MARC XML
    for element in root.iter():
        if not element.tag.endswith("datafield"):
            continue

        tag = element.get("tag", "")

        # Collect subfield values
        subfields_by_code = {}
        all_text_parts = []
        for child in element:
            if child.tag.endswith("subfield"):
                code = child.get("code", "")
                text = child.text or ""
                subfields_by_code.setdefault(code, []).append(text)
                all_text_parts.append(text)

        full_text = " ".join(all_text_parts)

        # Subject headings (650, 651, 600, 610, 630)
        if tag.startswith("6"):
            result["subjects"].append(full_text)

            # 650 $v Textbooks -- the MARC standard way to mark a textbook
            if tag == "650":
                for v_val in subfields_by_code.get("v", []):
                    if "textbook" in v_val.lower():
                        result["has_textbook_subject"] = True

        # Genre / Form (655)
        if tag == "655":
            result["genre_forms"].append(full_text)
            if "textbook" in full_text.lower():
                result["has_textbook_genre"] = True

        # Notes (500 general, 504 bibliography, 520 summary, 521 audience)
        if tag in ("500", "504", "520", "521"):
            result["notes"].append(full_text)

    return result


def extract_marc_from_bib(full_bib_json):
    """
    Pull and combine all MARC data from the 'anies' field in an Alma bib response.
    'anies' is a list that may contain one or more MARC XML strings.
    """
    combined = {
        "subjects": [],
        "genre_forms": [],
        "notes": [],
        "has_textbook_subject": False,
        "has_textbook_genre": False,
    }

    for item in full_bib_json.get("anies", []):
        marc_str = item if isinstance(item, str) else str(item)
        parsed = parse_marc_xml(marc_str)
        combined["subjects"].extend(parsed["subjects"])
        combined["genre_forms"].extend(parsed["genre_forms"])
        combined["notes"].extend(parsed["notes"])
        combined["has_textbook_subject"] = (
            combined["has_textbook_subject"] or parsed["has_textbook_subject"]
        )
        combined["has_textbook_genre"] = (
            combined["has_textbook_genre"] or parsed["has_textbook_genre"]
        )

    return combined

# ============================================================================
# TEXTBOOK SCORING
# ============================================================================

# Title patterns that are strong indicators of a textbook.
# Each entry: (regex pattern, human-readable label, point value)
TITLE_PATTERNS = [
    (r"\bintroduction\s+to\b",                    "Title: 'Introduction to'",              30),
    (r"\bfundamentals\s+of\b",                    "Title: 'Fundamentals of'",              30),
    (r"\bprinciples\s+of\b",                      "Title: 'Principles of'",                25),
    (r"\bessentials\s+of\b",                      "Title: 'Essentials of'",                25),
    (r"\bconcepts\s+(in|of)\b",                   "Title: 'Concepts in/of'",               25),
    (r"\bfoundations\s+of\b",                     "Title: 'Foundations of'",               25),
    (r"\boverview\s+of\b",                        "Title: 'Overview of'",                  20),
    (r"\bsurvey\s+of\b",                          "Title: 'Survey of'",                    20),
    (r"\bguide\s+to\b",                           "Title: 'Guide to'",                     15),
    (r"\bfor\s+(nurses?|nursing)\b",              "Title: 'for nurses/nursing'",           30),
    (r"\bfor\s+health\s+(care|professionals?)\b", "Title: 'for health professionals'",     30),
    (r"\bfor\s+(students?|beginners?)\b",         "Title: 'for students/beginners'",       20),
    (r"\b\d+(st|nd|rd|th)\s+ed(ition)?\b",        "Edition number in title",               20),
    (r"\b(workbook|study\s+guide)\b",             "Title: Workbook or Study Guide",        40),
    (r"\bmade\s+(incredibly\s+)?easy\b",          "Title: 'Made Easy' series",             35),
]

# Publishers that predominantly publish textbooks
ACADEMIC_PUBLISHERS = [
    "mcgraw-hill", "mcgraw hill", "pearson", "cengage", "wiley", "elsevier",
    "springer", "jones & bartlett", "jones and bartlett", "lippincott",
    "mosby", "wolters kluwer", "norton", "bedford", "worth publishers",
    "south-western", "thomson", "wadsworth", "brooks/cole", "brooks cole",
    "houghton mifflin", "addison-wesley", "addison wesley", "prentice hall",
    "saunders", "f.a. davis", "fa davis", "delmar", "routledge",
    "sage publications", "oxford university press", "cambridge university press",
]


def score_item(bib_data, marc_data):
    """
    Score one item for textbook likelihood.

    Args:
        bib_data  - dict from Alma brief bib API
                    (has: title, edition, publisher_const, ...)
        marc_data - dict from extract_marc_from_bib
                    (has: subjects, genre_forms, notes,
                          has_textbook_subject, has_textbook_genre)

    Returns:
        (confidence_level, score, reasons)
        confidence_level is "OBVIOUS", "LIKELY", or "UNCLEAR"
        score is an integer (higher = more textbook-like)
        reasons is a list of human-readable strings explaining the score
    """
    score = 0
    reasons = []

    title     = (bib_data.get("title")           or "").lower()
    edition   = (bib_data.get("edition")         or "").lower()
    publisher = (bib_data.get("publisher_const") or "").lower()

    # ------------------------------------------------------------------
    # OBVIOUS signals -- stop here, no need to accumulate more points
    # ------------------------------------------------------------------

    # Custom edition (written specifically for a course section)
    if re.search(r"\bcustom\s*(ed(ition)?|print|version)?\b", title, re.I) or \
       re.search(r"\bcustom\b", edition, re.I):
        return "OBVIOUS", 100, ["Custom edition (written for a specific course)"]

    # MARC subject heading: 650 $v Textbooks
    if marc_data.get("has_textbook_subject"):
        return "OBVIOUS", 100, [
            "Subject heading contains 'Textbooks' subdivision (650 $v Textbooks)"
        ]

    # MARC genre/form: Textbooks
    if marc_data.get("has_textbook_genre"):
        return "OBVIOUS", 100, ["Genre/Form field contains 'Textbooks' (655)"]

    # ------------------------------------------------------------------
    # LIKELY signals -- accumulate points
    # ------------------------------------------------------------------

    # Title pattern matches
    for pattern, label, pts in TITLE_PATTERNS:
        if re.search(pattern, title, re.I):
            score += pts
            reasons.append(label)

    # Has an edition statement (textbooks almost always list editions)
    if edition:
        score += 15
        reasons.append(f"Has edition statement: '{bib_data.get('edition')}'")

    # Academic / textbook publisher
    for pub in ACADEMIC_PUBLISHERS:
        if pub in publisher:
            score += 20
            reasons.append(f"Known textbook publisher: {bib_data.get('publisher_const')}")
            break

    # Notes field mentions course or textbook use
    notes_text = " ".join(marc_data.get("notes", [])).lower()
    if re.search(
        r"\b(textbook|for\s+use\s+in|course\s+(text|material)|undergraduate\s+course)\b",
        notes_text,
    ):
        score += 15
        reasons.append("Note field mentions course or textbook use")

    # Determine confidence level
    if score >= 50:
        return "LIKELY", score, reasons
    elif score > 0:
        return "UNCLEAR", score, reasons
    else:
        return "UNCLEAR", 0, ["No textbook indicators found"]


# ============================================================================
# FETCHING AND ENRICHMENT
# ============================================================================

def quick_score_passes(bib):
    """
    Fast check using brief bib data only (no MARC needed yet).
    Returns True if the item has any textbook signal worth investigating further.
    Items that fail this check are still included in the output as UNCLEAR.
    """
    title     = (bib.get("title")           or "").lower()
    edition   = (bib.get("edition")         or "")
    publisher = (bib.get("publisher_const") or "").lower()

    if re.search(r"\bcustom\b", title + " " + edition.lower(), re.I):
        return True
    if edition:
        return True
    for pattern, _, _ in TITLE_PATTERNS:
        if re.search(pattern, title, re.I):
            return True
    for pub in ACADEMIC_PUBLISHERS:
        if pub in publisher:
            return True
    return False


def fetch_all_stacks_bibs(config, limit=None):
    """
    Fetch brief bib records for all items in all configured Stacks locations.
    Deduplicates by MMS ID (a bib may appear in both STACK and OVERSIZE).

    Args:
        limit - if set, stop fetching after this many unique items (for test runs)
    """
    api_key   = config["alma_api_key"]
    locations = config.get("stacks_locations", ["STACK", "OVERSIZE"])
    delay     = config.get("api_delay_seconds", 0.3)

    seen = {}  # mms_id -> bib dict

    for location in locations:
        bibs = search_bibs_by_location(location, api_key, delay)
        for bib in bibs:
            mms_id = bib.get("mms_id")
            if not mms_id:
                continue
            if mms_id not in seen:
                seen[mms_id] = bib
            else:
                # Merge location label if same bib appears in multiple locations
                existing = seen[mms_id].get("_source_location", "")
                new_loc  = bib.get("_source_location", "")
                if new_loc and new_loc not in existing:
                    seen[mms_id]["_source_location"] = f"{existing}, {new_loc}"

        if limit and len(seen) >= limit:
            print(f"  --limit {limit} reached, stopping early.")
            break

    results = list(seen.values())
    if limit:
        results = results[:limit]
    return results


def load_bibs_from_csv(filepath, limit=None):
    """
    Load bib records from an Alma Analytics CSV export instead of the API.

    Expected columns (from a Physical Items report):
        MMS Id, Title, Author, Publisher, ISBN, Publication Date,
        Location Name, Subjects

    Returns a list of bib dicts in the same format used by the API path,
    so the rest of the script (enrichment, scoring, output) works unchanged.
    """
    import pandas as pd

    print(f"\n  Loading items from CSV: {filepath}")
    try:
        df = pd.read_csv(filepath, dtype=str).fillna("")
    except FileNotFoundError:
        print(f"  ERROR: File not found: {filepath}")
        sys.exit(1)

    print(f"  Rows in file: {len(df):,}")

    # Deduplicate by MMS Id — same bib can appear multiple times (one row per item)
    df = df.drop_duplicates(subset=["MMS Id"])
    print(f"  Unique MMS IDs after deduplication: {len(df):,}")

    if limit:
        df = df.head(limit)
        print(f"  *** TEST MODE: limiting to {limit} items ***")

    bibs = []
    for _, row in df.iterrows():
        # ISBN: Analytics exports multiple ISBNs as "; "-separated string
        isbn_raw = row.get("ISBN", "")
        isbn_list = [i.strip() for i in isbn_raw.split(";") if i.strip()] if isbn_raw else []

        bibs.append({
            "mms_id":           row.get("MMS Id", "").strip(),
            "title":            (row.get("Title (Complete)") or row.get("Title") or "").strip(),
            "author":           row.get("Author", "").strip(),
            "publisher_const":  row.get("Publisher", "").strip(),
            "isbn":             isbn_list,
            "date_of_publication": row.get("Publication Date", "").strip(),
            "edition":          "",   # not in Analytics export; filled by enrichment
            "_source_location": row.get("Location Name", "").strip(),
            "_subjects_csv":    row.get("Subjects", "").strip(),  # fallback if MARC unavailable
        })

    print(f"  Loaded {len(bibs):,} unique items.")
    return bibs


def enrich_items(bibs, config):
    """
    For each bib that passes the quick score, fetch:
      1. Full bib record (for MARC subject headings, genre/form, notes)
      2. Holdings record (for call number and confirmed location)

    Items that don't pass the quick score still get empty enrichment data
    so they appear in the output with an UNCLEAR rating.

    Returns the full list of bibs with '_marc_data', '_call_number',
    and '_confirmed_location' added to each.
    """
    api_key          = config["alma_api_key"]
    delay            = config.get("api_delay_seconds", 0.3)
    stacks_locations = set(config.get("stacks_locations", ["STACK", "OVERSIZE"]))

    to_enrich = [b for b in bibs if quick_score_passes(b)]
    to_skip   = [b for b in bibs if not quick_score_passes(b)]

    print(f"\n  {len(to_enrich):,} items flagged for full MARC + holdings fetch")
    print(f"  {len(to_skip):,} items skipped (no textbook signals in brief data)")

    enriched = []
    for i, bib in enumerate(to_enrich, 1):
        mms_id = bib.get("mms_id")
        if not mms_id:
            continue

        if i % 25 == 0 or i == len(to_enrich):
            print(f"  Enriching {i:,} / {len(to_enrich):,} ...", end="\r")

        # Full bib -> MARC data
        full_bib = get_full_bib(mms_id, api_key)
        bib["_marc_data"] = extract_marc_from_bib(full_bib) if full_bib else {}

        # Holdings -> call number + confirmed location
        holdings = get_holdings(mms_id, api_key)
        stacks_h = [
            h for h in holdings
            if h.get("location", {}).get("value", "") in stacks_locations
        ]
        bib["_holdings"]           = stacks_h
        bib["_call_number"]        = stacks_h[0].get("call_number", "") if stacks_h else ""
        bib["_confirmed_location"] = (
            stacks_h[0].get("location", {}).get("value", "")
            if stacks_h else bib.get("_source_location", "")
        )

        enriched.append(bib)
        time.sleep(delay)

    print(f"\n  Enrichment complete.")

    # Give skipped items empty enrichment stubs
    for bib in to_skip:
        bib["_marc_data"]          = {}
        bib["_holdings"]           = []
        bib["_call_number"]        = ""
        bib["_confirmed_location"] = bib.get("_source_location", "")

    return enriched + to_skip

# ============================================================================
# BUILD DATAFRAME
# ============================================================================

def build_dataframe(bibs, config):
    """
    Convert the enriched bib list into a Pandas DataFrame.
    Scores each item and sorts results by confidence level.
    """
    rows = []
    for bib in bibs:
        marc = bib.get("_marc_data", {})
        confidence, conf_score, reasons = score_item(bib, marc)

        isbn_raw = bib.get("isbn", [])
        isbn_str = (
            "; ".join(isbn_raw) if isinstance(isbn_raw, list) else str(isbn_raw or "")
        )

        rows.append({
            "MMS_ID":             bib.get("mms_id", ""),
            "Title":              bib.get("title", ""),
            "Author":             bib.get("author", ""),
            "Call_Number":        bib.get("_call_number", ""),
            "Location":           bib.get("_confirmed_location", bib.get("_source_location", "")),
            "Edition":            bib.get("edition", ""),
            "Publisher":          bib.get("publisher_const", ""),
            "Pub_Year":           bib.get("date_of_publication", ""),
            "ISBN":               isbn_str,
            "Confidence":         confidence,
            "Confidence_Score":   conf_score,
            "Confidence_Reasons": "; ".join(reasons),
            "Subjects":           ("; ".join(marc.get("subjects", [])))[:500],
        })

    df = pd.DataFrame(rows)

    if df.empty:
        print("  WARNING: No items to analyze.")
        return df

    # Map confidence to a plain-English recommended action
    action_map = {
        "OBVIOUS": "MOVE TO OLD TEXTBOOKS",
        "LIKELY":  "LIKELY MOVE - Confirm it's a textbook",
        "UNCLEAR": "REVIEW MANUALLY",
    }
    df["Recommended_Action"] = df["Confidence"].map(action_map)

    # Sort: OBVIOUS first, then LIKELY, then UNCLEAR, then by title
    conf_order = {"OBVIOUS": 0, "LIKELY": 1, "UNCLEAR": 2}
    df["_co"] = df["Confidence"].map(conf_order).fillna(99)
    df = (
        df.sort_values(["_co", "Title"])
          .drop(columns=["_co"])
          .reset_index(drop=True)
    )

    return df

# ============================================================================
# EXCEL OUTPUT
# ============================================================================

CONF_COLORS = {
    "OBVIOUS": "FFB3B3",  # soft red
    "LIKELY":  "FFE0B3",  # soft orange
    "UNCLEAR": "FFFDB3",  # soft yellow
}

COLUMN_WIDTHS = {
    "MMS_ID": 18, "Title": 50, "Author": 25, "Call_Number": 18,
    "Location": 12, "Edition": 22,
    "Publisher": 22, "Pub_Year": 8, "ISBN": 22,
    "Confidence": 12, "Confidence_Score": 8,
    "Confidence_Reasons": 45,
    "Recommended_Action": 32,
    "Subjects": 40,
    "Decision": 22, "Decision_Notes": 30,
}

OUTPUT_COLUMNS = [
    "MMS_ID", "Title", "Author", "Call_Number", "Location",
    "Edition", "Publisher", "Pub_Year", "ISBN",
    "Confidence", "Confidence_Score", "Confidence_Reasons",
    "Recommended_Action",
    "Subjects",
    "Decision",        # blank -- reviewer fills this in
    "Decision_Notes",  # blank -- reviewer notes
]


def save_excel(df, output_path):
    """
    Write the analysis DataFrame to an Excel file with:
    - Color-coded rows by confidence level (red/orange/yellow)
    - Bold, wrapped header row
    - Frozen top row
    - Fixed column widths
    - Blank Decision and Decision_Notes columns for manual review
    """
    df = df.copy()
    df["Decision"]       = ""
    df["Decision_Notes"] = ""

    cols   = [c for c in OUTPUT_COLUMNS if c in df.columns]
    df_out = df[cols]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Stacks Analysis")
        ws = writer.sheets["Stacks Analysis"]

        conf_col_idx = cols.index("Confidence") + 1  # 1-based column index

        # Color each data row by confidence level
        for row_num in range(2, len(df_out) + 2):
            conf_val  = ws.cell(row=row_num, column=conf_col_idx).value
            hex_color = CONF_COLORS.get(str(conf_val), "FFFFFF")
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            for col_num in range(1, len(cols) + 1):
                ws.cell(row=row_num, column=col_num).fill = fill

        # Format header row
        for col_num, col_name in enumerate(cols, 1):
            cell           = ws.cell(row=1, column=col_num)
            cell.font      = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True)

        # Set column widths
        for col_num, col_name in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = COLUMN_WIDTHS.get(col_name, 15)

        ws.freeze_panes = "A2"

    print(f"  Saved: {output_path}")

# ============================================================================
# CACHE HELPERS
# ============================================================================

def save_cache(data, path):
    path = Path(path)
    path.parent.mkdir(exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f)
    print(f"  Cached {len(data):,} items -> {path.name}")


def load_cache(path):
    path = Path(path)
    if not path.exists():
        return None
    with open(path) as f:
        data = json.load(f)
    print(f"  Loaded {len(data):,} items from cache: {path.name}")
    return data

# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Analyze Stacks for textbooks that should move to Old Textbooks",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python analyze.py              # fetch fresh data from Alma + run analysis
  python analyze.py --cached     # skip API fetch, reanalyze from last cached data
  python analyze.py --fetch-only # fetch and cache data, stop before analysis
        """,
    )
    parser.add_argument(
        "--cached",
        action="store_true",
        help="Use cached data from a previous run (no API calls)",
    )
    parser.add_argument(
        "--fetch-only",
        action="store_true",
        help="Fetch and cache data, then exit without building the spreadsheet",
    )
    parser.add_argument(
        "--sandbox",
        action="store_true",
        help="Use sandbox API key and config (config_sandbox.json). Safe for testing.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        metavar="N",
        help="Stop after fetching/loading N items (e.g. --limit 100 for a quick test run).",
    )
    parser.add_argument(
        "--input",
        type=str,
        default=None,
        metavar="FILE",
        help="Path to an Alma Analytics CSV export. Skips the Alma API fetch step.",
    )
    args = parser.parse_args()

    print("=" * 65)
    print("OLD TEXTBOOKS STACKS ANALYZER")
    if args.sandbox:
        print("*** SANDBOX MODE — using sandbox API, no production data affected ***")
    print("=" * 65)

    config  = load_config(sandbox=args.sandbox)
    api_key = config.get("alma_api_key", "")
    if not api_key:
        print("ERROR: No alma_api_key found in config.json")
        sys.exit(1)

    # Use a separate cache folder for sandbox so production data is never overwritten
    cache_dir      = CACHE_DIR / ("sandbox" if args.sandbox else "production")
    cache_dir.mkdir(parents=True, exist_ok=True)
    raw_cache      = cache_dir / "stacks_raw.json"
    enriched_cache = cache_dir / "stacks_enriched.json"

    # ---- Step 1: Get brief bib data ----
    if args.cached and enriched_cache.exists():
        print("\n[Step 1+2] Loading enriched data from cache (skipping API)...")
        bibs = load_cache(enriched_cache)

    elif args.cached and raw_cache.exists():
        print("\n[Step 1] Loading brief bib data from cache...")
        bibs = load_cache(raw_cache)
        print("\n[Step 2] Enriching with full MARC + holdings...")
        bibs = enrich_items(bibs, config)
        save_cache(bibs, enriched_cache)

    elif args.input:
        # CSV export from Alma Analytics -- no API search needed
        print("\n[Step 1] Loading items from Alma Analytics CSV export...")
        bibs = load_bibs_from_csv(args.input, limit=args.limit)
        print(f"\n  Total unique bibs: {len(bibs):,}")
        save_cache(bibs, raw_cache)

        print("\n[Step 2] Enriching with full MARC + holdings data from Alma API...")
        bibs = enrich_items(bibs, config)
        save_cache(bibs, enriched_cache)

    else:
        print("\n[Step 1] Fetching brief bib data from Alma API...")
        if args.limit:
            print(f"  *** TEST MODE: limiting to {args.limit} items ***")
        bibs = fetch_all_stacks_bibs(config, limit=args.limit)
        print(f"\n  Total unique bibs: {len(bibs):,}")
        save_cache(bibs, raw_cache)

        if args.fetch_only:
            print("\nFetch complete. Run again without --fetch-only to analyze.")
            return

        print("\n[Step 2] Enriching with full MARC + holdings data...")
        bibs = enrich_items(bibs, config)
        save_cache(bibs, enriched_cache)

    if not bibs:
        print("\nNo items found. Check your location codes in config.json.")
        return

    # ---- Step 3: Analyze ----
    print(f"\n[Step 3] Scoring {len(bibs):,} items...")
    df = build_dataframe(bibs, config)

    if df.empty:
        print("No items to save.")
        return

    # ---- Step 4: Save Excel ----
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"stacks_analysis_{timestamp}.xlsx"
    latest_path = OUTPUT_DIR / "latest_analysis.xlsx"

    print(f"\n[Step 4] Saving spreadsheet...")
    save_excel(df, output_path)
    save_excel(df, latest_path)

    # ---- Summary ----
    print("\n" + "=" * 65)
    print("SUMMARY")
    print("=" * 65)
    total   = len(df)
    obvious = (df["Confidence"] == "OBVIOUS").sum()
    likely  = (df["Confidence"] == "LIKELY").sum()
    unclear = (df["Confidence"] == "UNCLEAR").sum()

    print(f"Total items analyzed:        {total:,}")
    print(f"  OBVIOUS textbooks:         {obvious:,}")
    print(f"  LIKELY textbooks:          {likely:,}")
    print(f"  UNCLEAR (review manually): {unclear:,}")
    print()
    print(f"Spreadsheet:  {output_path}")
    print(f"Web app file: {latest_path}")
    print()
    print("Next step: run   python review_app.py   to open the review interface")
    print("=" * 65)


if __name__ == "__main__":
    main()
