"""
export_barcodes.py
Reads 101425_MissingItemsReport_separated.xlsx and exports just the barcodes
from two sheets into separate Excel files for further system processing:

  - barcodes_not_found_with_notes.xlsx
  - barcodes_not_found_no_notes.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

INPUT_FILE = "101425_MissingItemsReport_separated.xlsx"
BARCODE_COL = 6  # Column F

EXPORTS = [
    ("Not Found With Notes", "barcodes_not_found_with_notes.xlsx", "FFC7CE"),  # red
    ("Not Found No Notes",   "barcodes_not_found_no_notes.xlsx",   "F2CEEF"),  # purple
]


def main():
    input_path = Path(INPUT_FILE)
    print(f"Loading: {input_path}\n")
    wb_in = openpyxl.load_workbook(input_path)

    for sheet_name, output_filename, header_color in EXPORTS:
        ws_in = wb_in[sheet_name]

        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "Barcodes"

        # Write header
        ws_out.append(["Barcode"])
        header_cell = ws_out["A1"]
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill("solid", fgColor=header_color)
        header_cell.alignment = Alignment(horizontal="center")

        # Write barcodes (skip header row)
        count = 0
        for row in ws_in.iter_rows(min_row=2, values_only=True):
            barcode = row[BARCODE_COL - 1]
            if barcode is not None:
                ws_out.append([barcode])
                count += 1

        # Widen the barcode column
        ws_out.column_dimensions["A"].width = 20

        output_path = Path(output_filename)
        wb_out.save(output_path)
        print(f"  {sheet_name}: {count} barcodes → {output_filename}")

    print("\nDone.")


if __name__ == "__main__":
    main()
