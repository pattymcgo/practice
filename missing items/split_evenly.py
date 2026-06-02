"""
split_evenly.py
Reads a missing items report and splits rows evenly (in order) among three
staff members, rather than splitting by location.

Input:  'June2026_Missing Items Report.xlsx'
Output: 'June2026_Missing_Items_split.xlsx'

  IC - first third of rows
  PL - second third
  YH - final third

(Rows are kept in their original order within each sheet.)

WHEN TO USE THIS SCRIPT:
  Use when you want to divide work evenly among 3 staff and don't care about
  splitting by location. Good for smaller reports (~100-150 items).

TO REUSE FOR A NEW FILE:
  1. Update INPUT_FILE and OUTPUT_FILE near the top.
  2. If headers are on a different row, update HEADER_ROW and DATA_START.
  3. If you want different sheet names or colors, update OUTPUT_SHEETS.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

INPUT_FILE  = "June2026_Missing Items Report.xlsx"
OUTPUT_FILE = "June2026_Missing_Items_split.xlsx"

HEADER_ROW = 3   # real column headers are on row 3
DATA_START  = 4  # data begins on row 4

OUTPUT_SHEETS = [
    ("IC", "BDD7EE"),   # blue
    ("PL", "FFC7CE"),   # red
    ("YH", "C6EFCE"),   # green
]


def style_header_row(ws, hex_color):
    """Apply bold text and background fill to the header row."""
    fill = PatternFill("solid", fgColor=hex_color)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font      = bold
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center")


def autosize_columns(ws):
    """Set column widths based on content (capped at 50)."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def main():
    input_path  = Path(INPUT_FILE)
    output_path = Path(OUTPUT_FILE)

    print(f"Loading: {input_path}")
    wb_in = openpyxl.load_workbook(input_path)
    ws_in = wb_in["Sheet1"]

    # Read real headers from row 3
    headers = [cell.value for cell in ws_in[HEADER_ROW]]

    # Collect all non-empty data rows
    all_rows = []
    for row_idx in range(DATA_START, ws_in.max_row + 1):
        row_values = [
            ws_in.cell(row=row_idx, column=c).value
            for c in range(1, ws_in.max_column + 1)
        ]
        if any(v is not None for v in row_values):
            all_rows.append(row_values)

    total = len(all_rows)
    print(f"Total data rows: {total}")

    # Calculate cutoffs for three roughly equal groups
    # e.g. 137 rows → 46, 46, 45
    base      = total // 3
    remainder = total % 3
    # First 'remainder' groups get one extra row so sizes differ by at most 1
    sizes = [base + (1 if i < remainder else 0) for i in range(3)]
    cutoffs = [sum(sizes[:i]) for i in range(4)]  # start indices for each group

    # Set up output workbook
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    out_sheets = {}
    for sheet_name, color in OUTPUT_SHEETS:
        ws = wb_out.create_sheet(title=sheet_name)
        ws.append(headers)
        style_header_row(ws, color)
        out_sheets[sheet_name] = ws

    # Distribute rows evenly
    for i, (sheet_name, _) in enumerate(OUTPUT_SHEETS):
        group_rows = all_rows[cutoffs[i]:cutoffs[i + 1]]
        for row in group_rows:
            out_sheets[sheet_name].append(row)

    # Auto-size columns
    for ws in out_sheets.values():
        autosize_columns(ws)

    wb_out.save(output_path)

    print("\nResults:")
    for i, (sheet_name, _) in enumerate(OUTPUT_SHEETS):
        count = cutoffs[i + 1] - cutoffs[i]
        print(f"  {sheet_name}: {count} rows")
    print(f"\nSaved: {output_path}")


if __name__ == "__main__":
    main()
