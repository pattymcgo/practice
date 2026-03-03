"""
split_missing_report.py
Reads 'Feb2026 Missing Items Report.xlsx' (headers on row 3, data from row 4)
and splits all 238 rows into three sheets as evenly as possible by location:

  IC  (79)  - Stacks, first 79 rows
  PL  (79)  - Stacks, remaining 59 rows + Comics + Juvenile + Textbooks + Dictionaries
  YH  (80)  - Reserves, 3 hours + Leisure Reading

Output: Feb2026_Missing_Items_split.xlsx

WHEN TO USE THIS SCRIPT:
  Use this script when your file is a raw system export with only Sheet1,
  no Found? values filled in, and a title row above the headers
  (headers on row 3, data starting on row 4).
  Do NOT run clean_missing_items.py or separate_by_status.py on these files.

TO REUSE FOR A NEW FILE:
  1. Update INPUT_FILE and OUTPUT_FILE near the top of the script.
  2. Check the location counts in the new file (run a quick inspection first)
     and update YH_LOCATIONS, PL_LOCATIONS, and STACKS_IC_LIMIT to reflect
     the new distribution and achieve the closest split to 1/3 each.
  3. If headers are on a different row, update HEADER_ROW and DATA_START.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

INPUT_FILE  = "Feb2026 Missing Items Report.xlsx"
OUTPUT_FILE = "Feb2026_Missing_Items_split.xlsx"

HEADER_ROW   = 3  # real column headers are on row 3
DATA_START   = 4  # data begins on row 4
LOCATION_COL = 2  # Column B = Location Name

# Locations assigned to each sheet (whole groups kept together)
YH_LOCATIONS = {"Reserves, 3 hours", "Leisure Reading"}
PL_LOCATIONS = {"Comics and Graphic Novels Collection", "Juvenile Collection",
                "Textbooks", "Dictionaries"}
# Stacks is split: first 79 rows → IC, remainder → PL
STACKS_LOCATION = "Stacks"
STACKS_IC_LIMIT = 79  # how many Stacks rows go to IC before overflow to PL

OUTPUT_SHEETS = [
    ("IC", "BDD7EE"),   # blue
    ("PL", "FFC7CE"),   # red
    ("YH", "C6EFCE"),   # green
]


def style_header_row(ws, hex_color):
    """Apply bold text and background fill to the header row."""
    fill = PatternFill("solid", fgColor=hex_color)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font      = bold
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center")


def autosize_columns(ws):
    """Set column widths based on content (capped at 50)."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def main():
    input_path  = Path(INPUT_FILE)
    output_path = Path(OUTPUT_FILE)

    print(f"Loading: {input_path}")
    wb_in = openpyxl.load_workbook(input_path)
    ws_in = wb_in["Sheet1"]

    # Read real headers from row 3
    headers = [cell.value for cell in ws_in[HEADER_ROW]]

    # Set up output workbook
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # remove default blank sheet

    out_sheets = {}
    for sheet_name, color in OUTPUT_SHEETS:
        ws = wb_out.create_sheet(title=sheet_name)
        ws.append(headers)
        style_header_row(ws, color)
        out_sheets[sheet_name] = ws

    counts       = {name: 0 for name, _ in OUTPUT_SHEETS}
    stacks_to_ic = 0  # tracks how many Stacks rows have gone to IC

    for row_idx in range(DATA_START, ws_in.max_row + 1):
        row_values = [
            ws_in.cell(row=row_idx, column=c).value
            for c in range(1, ws_in.max_column + 1)
        ]

        # Skip fully empty rows
        if all(v is None for v in row_values):
            continue

        location = row_values[LOCATION_COL - 1]

        if location == STACKS_LOCATION:
            if stacks_to_ic < STACKS_IC_LIMIT:
                target = "IC"
                stacks_to_ic += 1
            else:
                target = "PL"

        elif location in YH_LOCATIONS:
            target = "YH"

        elif location in PL_LOCATIONS:
            target = "PL"

        else:
            # Unexpected location — flag it and send to YH
            print(f"  WARNING: unknown location '{location}' on row {row_idx} → assigned to YH")
            target = "YH"

        out_sheets[target].append(row_values)
        counts[target] += 1

    # Auto-size columns
    for ws in out_sheets.values():
        autosize_columns(ws)

    wb_out.save(output_path)

    print("\nResults:")
    for sheet_name, _ in OUTPUT_SHEETS:
        print(f"  {sheet_name}: {counts[sheet_name]} rows")
    print(f"\nSaved: {output_path}")


if __name__ == "__main__":
    main()
