"""
clean_missing_items.py
Cleans 10152025_LostItemsReport.xlsx:
  - Normalizes 'Found?' values (consistent casing, strips whitespace)
  - Flags blank 'Found?' rows in YH as 'Needs review'
  - Populates Sheet1 'Found?' by looking up barcodes from IC/YH/PL sub-sheets
  - Removes fully empty rows from Sheet1
  - Saves result to 10152025_LostItemsReport_cleaned.xlsx

WHEN TO USE THIS SCRIPT:
  Use this script when your file has Sheet1 PLUS sub-sheets (IC, YH, PL).
  It uses the sub-sheets to fill in the Found? column on Sheet1.
  Run this first, then run separate_by_status.py on the cleaned output.

  Skip this script if your file only has Sheet1 — go straight to
  separate_by_status.py (if Found? values are filled in) or
  split_missing_report.py (if Found? values are blank).

TO REUSE FOR A NEW FILE:
  Update INPUT_FILE and OUTPUT_FILE near the top of the script.
"""

import openpyxl
from copy import copy
from pathlib import Path

INPUT_FILE = "10152025_LostItemsReport.xlsx"
OUTPUT_FILE = "10152025_LostItemsReport_cleaned.xlsx"

FOUND_COL = 1   # Column A (1-indexed)
BARCODE_COL = 7  # Column G


def normalize_found(value):
    """Standardize Found? values to canonical forms."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return None  # treat whitespace-only as missing
        if stripped.lower() == "not found":
            return "Not found"
        if stripped.lower() == "found":
            return "Found"
        return stripped  # preserve anything else (e.g. 'Needs review')
    return value


def build_barcode_lookup(wb, sheet_names):
    """Build a dict of barcode -> Found? from the given sub-sheets."""
    lookup = {}
    for name in sheet_names:
        ws = wb[name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            barcode = row[BARCODE_COL - 1]
            found = normalize_found(row[FOUND_COL - 1])
            if barcode and found is not None:
                # Don't overwrite a 'Found' entry with 'Not found'
                if barcode not in lookup or found == "Found":
                    lookup[barcode] = found
    return lookup


def clean_sheet(ws, sheet_name, barcode_lookup=None):
    """
    Clean a single worksheet in place.
    - Normalize Found? values
    - If barcode_lookup provided, fill missing Found? from it
    - Flag remaining None Found? as 'Needs review' (for YH)
    - Remove fully empty rows
    Returns number of rows removed.
    """
    rows_to_delete = []

    for row_idx in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]

        # Mark fully empty rows for deletion
        if all(v is None for v in row_values):
            rows_to_delete.append(row_idx)
            continue

        found_cell = ws.cell(row=row_idx, column=FOUND_COL)
        barcode_cell = ws.cell(row=row_idx, column=BARCODE_COL)

        normalized = normalize_found(found_cell.value)

        if normalized is None:
            # Try to fill from lookup (Sheet1 uses this)
            if barcode_lookup and barcode_cell.value:
                normalized = barcode_lookup.get(str(barcode_cell.value))

            # Still None: flag it
            if normalized is None:
                normalized = "Needs review"

        found_cell.value = normalized

    # Delete empty rows bottom-up to preserve indices
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)

    return len(rows_to_delete)


def main():
    input_path = Path(INPUT_FILE)
    output_path = Path(OUTPUT_FILE)

    print(f"Loading: {input_path}")
    wb = openpyxl.load_workbook(input_path)

    sub_sheets = ["IC", "YH", "PL"]

    # Step 1: Build barcode -> Found? lookup from sub-sheets (before cleaning them)
    print("Building barcode lookup from IC / YH / PL ...")
    barcode_lookup = build_barcode_lookup(wb, sub_sheets)
    print(f"  {len(barcode_lookup)} unique barcodes indexed")

    # Step 2: Clean sub-sheets (normalize values, flag blanks)
    for name in sub_sheets:
        ws = wb[name]
        removed = clean_sheet(ws, name)
        print(f"  {name}: cleaned (removed {removed} empty rows)")

    # Step 3: Clean Sheet1 — populate Found? from lookup
    ws1 = wb["Sheet1"]
    removed = clean_sheet(ws1, "Sheet1", barcode_lookup=barcode_lookup)
    print(f"  Sheet1: cleaned (removed {removed} empty rows)")

    # Step 4: Report any Sheet1 rows that couldn't be matched
    unmatched = 0
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if row[FOUND_COL - 1] == "Needs review":
            unmatched += 1
    if unmatched:
        print(f"  WARNING: {unmatched} Sheet1 rows could not be matched to a sub-sheet — marked 'Needs review'")

    # Step 5: Save
    wb.save(output_path)
    print(f"\nSaved cleaned file: {output_path}")


if __name__ == "__main__":
    main()
