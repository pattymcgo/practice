"""
separate_by_status.py
Reads Sheet1 from 101425_MissingItemsReport_cleaned.xlsx and separates rows
into four categories based on Found? status and notes content:

  1. Found Items           - Found? is any 'found' variant
  2. Needs Review          - Found? is 'Needs review', OR call number is 'Unknown'
  3. Not Found With Notes  - Not found AND Fulfillment Note or Internal Note 2
                             contains 'last searched' (extra spaces ok), 'discard',
                             'not found on shelf', 'not on shelf', 'wanted to weed',
                             or 'missing'
  4. Not Found No Notes    - Not found AND neither note column contains those keywords

Output: 101425_MissingItemsReport_separated.xlsx

WHEN TO USE THIS SCRIPT:
  Use this script when your file has Sheet1 with Found? values already filled in
  (headers on row 1, data from row 2).
  - If the file also has sub-sheets (IC, YH, PL): run clean_missing_items.py first
    to populate Found? values, then run this script on the cleaned output.
  - If the file only has Sheet1 with Found? already filled in: run this script
    directly (no need for clean_missing_items.py).
  - If Found? values are blank (raw system export): use split_missing_report.py
    instead.

TO REUSE FOR A NEW FILE:
  Update INPUT_FILE and OUTPUT_FILE near the top of the script.
  If your notes are in different columns, update FULFILLMENT_NOTE_COL and
  INTERNAL_NOTE_COL. Add new keywords to NOTE_PATTERNS as needed.
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

INPUT_FILE  = "101425_MissingItemsReport_cleaned.xlsx"
OUTPUT_FILE = "101425_MissingItemsReport_separated.xlsx"

FOUND_COL            = 1   # Column A
CALL_NUMBER_COL      = 4   # Column D
FULFILLMENT_NOTE_COL = 9   # Column I
INTERNAL_NOTE_COL    = 10  # Column J

# Regex patterns to match in notes columns (case-insensitive).
# \s+ catches single or multiple spaces between words.
NOTE_PATTERNS = [
    re.compile(r"last\s+searched",         re.IGNORECASE),
    re.compile(r"discard",                 re.IGNORECASE),
    re.compile(r"not\s+found\s+on\s+shelf", re.IGNORECASE),
    re.compile(r"not\s+on\s+shelf",        re.IGNORECASE),
    re.compile(r"wanted\s+to\s+weed",      re.IGNORECASE),
    re.compile(r"missing",                 re.IGNORECASE),
]

# Found? values that count as "found"
FOUND_VALUES = {
    "found",
    "found, removed missing status",
    "found, missing status removed",
}

# Found? values that count as "not found"
NOT_FOUND_VALUES = {"not found", "no"}

# Output sheet names and their header colors
OUTPUT_SHEETS = [
    ("Found Items",          "C6EFCE"),  # green
    ("Needs Review",         "FFEB9C"),  # yellow
    ("Not Found With Notes", "FFC7CE"),  # red
    ("Not Found No Notes",   "F2CEEF"),  # purple
]


def get_status(found_val):
    """Return canonical status string from a Found? cell value."""
    if found_val is None:
        return "needs review"
    return str(found_val).strip().lower()


def notes_contain_keywords(row_values):
    """Return True if either notes column matches any of NOTE_PATTERNS."""
    for col in [FULFILLMENT_NOTE_COL, INTERNAL_NOTE_COL]:
        cell_val = row_values[col - 1]
        if cell_val:
            text = str(cell_val)
            if any(p.search(text) for p in NOTE_PATTERNS):
                return True
    return False


def style_header_row(ws, hex_color):
    """Apply bold text and background fill to the header row."""
    fill = PatternFill("solid", fgColor=hex_color)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font  = bold
        cell.fill  = fill
        cell.alignment = Alignment(horizontal="center")


def autosize_columns(ws):
    """Set column widths based on content (capped at 50)."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def main():
    input_path  = Path(INPUT_FILE)
    output_path = Path(OUTPUT_FILE)

    print(f"Loading: {input_path}")
    wb_in = openpyxl.load_workbook(input_path)
    ws_in = wb_in["Sheet1"]

    # Read headers from Sheet1
    headers = [cell.value for cell in ws_in[1]]

    # Set up output workbook with one sheet per category
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # remove default blank sheet

    out_sheets = {}
    for sheet_name, color in OUTPUT_SHEETS:
        ws = wb_out.create_sheet(title=sheet_name)
        ws.append(headers)         # write header row
        style_header_row(ws, color)
        out_sheets[sheet_name] = ws

    counts = {name: 0 for name, _ in OUTPUT_SHEETS}

    # Process each data row in Sheet1
    for row_idx in range(2, ws_in.max_row + 1):
        row_values = [
            ws_in.cell(row=row_idx, column=c).value
            for c in range(1, ws_in.max_column + 1)
        ]

        # Skip fully empty rows
        if all(v is None for v in row_values):
            continue

        call_number = row_values[CALL_NUMBER_COL - 1]
        status      = get_status(row_values[FOUND_COL - 1])

        # Unknown call number takes priority — flag for review regardless of Found? status
        if call_number and str(call_number).strip().lower() == "unknown":
            target = "Needs Review"

        elif status in FOUND_VALUES:
            target = "Found Items"

        elif status == "needs review":
            target = "Needs Review"

        elif status in NOT_FOUND_VALUES:
            # Check notes to decide which not-found bucket
            if notes_contain_keywords(row_values):
                target = "Not Found With Notes"
            else:
                target = "Not Found No Notes"

        else:
            # Unexpected value — send to Needs Review for safety
            target = "Needs Review"

        out_sheets[target].append(row_values)
        counts[target] += 1

    # Auto-size columns in every output sheet
    for ws in out_sheets.values():
        autosize_columns(ws)

    wb_out.save(output_path)

    print("\nResults:")
    for sheet_name, _ in OUTPUT_SHEETS:
        print(f"  {sheet_name:25s}: {counts[sheet_name]} rows")
    print(f"\nSaved: {output_path}")


if __name__ == "__main__":
    main()
