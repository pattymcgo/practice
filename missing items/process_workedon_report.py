"""
process_workedon_report.py
Reads Feb2026_Missing_Items_workedon.xlsx (sheets: IC, PL, YH),
merges all rows, and writes a processed workbook with four sheets:

  1. Not Found - Barcodes    : barcode column only, all not-found items
  2. Found Items             : all columns, items marked found
  3. Not Found - With Notes  : all columns, not found + keyword in notes
  4. Not Found - No Notes    : all columns, not found + no keyword in notes

TO REUSE FOR A NEW FILE:
  Update INPUT_FILE and OUTPUT_FILE at the top.
  If staff used different Found? spellings, add them to IS_FOUND or IS_NOT_FOUND.
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

INPUT_FILE  = "Feb2026_Missing_Items_workedon.xlsx"
OUTPUT_FILE = "Feb2026_Missing_Items_processed.xlsx"

# Source sheets to read (in order)
SOURCE_SHEETS = ["IC", "PL", "YH"]

# Column index (0-based) for key fields — based on current file layout
FOUND_COL   = 0   # A: Found?
BARCODE_COL = 8   # I: Barcode
NOTE_COLS   = [11, 12, 13, 14]  # L-O: Internal Note 1-3, Fulfillment Note

# Found? values that mean "found" (lowercased, stripped)
IS_FOUND = {"yes", "found", "on loan"}

# Found? values that mean "not found" (lowercased, stripped)
IS_NOT_FOUND = {"nf", "no", "n", "not found"}

# Keywords in notes that flag extra context
NOTE_PATTERNS = [
    re.compile(r"last\s+searched",          re.IGNORECASE),
    re.compile(r"discard",                  re.IGNORECASE),
    re.compile(r"not\s+found\s+on\s+shelf", re.IGNORECASE),
    re.compile(r"not\s+on\s+shelf",         re.IGNORECASE),
    re.compile(r"wanted\s+to\s+weed",       re.IGNORECASE),
    re.compile(r"missing",                  re.IGNORECASE),
]

# Output sheets: (name, header colour)
OUTPUT_SHEETS = [
    ("Not Found - Barcodes",   "F2CEEF"),  # purple
    ("Found Items",            "C6EFCE"),  # green
    ("Not Found - With Notes", "FFC7CE"),  # red
    ("Not Found - No Notes",   "FFEB9C"),  # yellow
]


def is_found(val):
    """Return True if Found? value is any 'found' variant."""
    if val is None:
        return False
    # Starts-with check catches "Yes - Stacks", "Yes - Reserves", etc.
    v = str(val).strip().lower()
    return v.startswith("yes") or v in IS_FOUND


def is_not_found(val):
    """Return True if Found? value is a 'not found' variant."""
    if val is None:
        return False
    return str(val).strip().lower() in IS_NOT_FOUND


def has_note_content(row):
    """Return True if any notes column has any non-empty content."""
    for col in NOTE_COLS:
        if col < len(row) and row[col] and str(row[col]).strip():
            return True
    return False


def style_header(ws, hex_color):
    fill = PatternFill("solid", fgColor=hex_color)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font      = bold
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center")


def autosize(ws):
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def main():
    input_path  = Path(INPUT_FILE)
    output_path = Path(OUTPUT_FILE)

    print(f"Loading: {input_path}")
    wb_in = openpyxl.load_workbook(input_path)

    # Read headers from first source sheet (all sheets share the same layout)
    headers = [cell.value for cell in wb_in[SOURCE_SHEETS[0]][1]]

    # Build output workbook
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    out = {}
    for name, color in OUTPUT_SHEETS:
        ws = wb_out.create_sheet(title=name)
        if name == "Not Found - Barcodes":
            ws.append(["Barcode"])
        else:
            ws.append(headers)
        style_header(ws, color)
        out[name] = ws

    counts = {name: 0 for name, _ in OUTPUT_SHEETS}
    skipped = 0

    for sheet_name in SOURCE_SHEETS:
        ws_in = wb_in[sheet_name]
        for row in ws_in.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue

            found_val = row[FOUND_COL]

            if is_found(found_val):
                out["Found Items"].append(list(row))
                counts["Found Items"] += 1

            elif is_not_found(found_val):
                barcode = row[BARCODE_COL]
                if barcode is not None:
                    out["Not Found - Barcodes"].append([barcode])
                    counts["Not Found - Barcodes"] += 1

                if has_note_content(row):
                    out["Not Found - With Notes"].append(list(row))
                    counts["Not Found - With Notes"] += 1
                else:
                    out["Not Found - No Notes"].append(list(row))
                    counts["Not Found - No Notes"] += 1

            else:
                # Unexpected value — report it but don't lose the row
                print(f"  WARNING: unrecognised Found? value '{found_val}' — skipped")
                skipped += 1

    for ws in out.values():
        autosize(ws)

    wb_out.save(output_path)

    print("\nResults:")
    for name, _ in OUTPUT_SHEETS:
        print(f"  {name:28s}: {counts[name]} rows")
    if skipped:
        print(f"  Skipped (unrecognised):      {skipped} rows")
    print(f"\nSaved: {output_path}")


if __name__ == "__main__":
    main()
