"""
Create Staff Review Workbook
=============================
Builds a shared Excel workbook with:
  - Data          : full Primo results
  - AJ/AM/GM/NP/PL/RW : data divided evenly among staff for review
  - LW - Purchase : blank purchase request sheet
  - TS - Put on Reserves   : blank put-on-reserves sheet
  - TS - Take off Reserves : blank take-off-reserves sheet

Usage:
    python create_staff_workbook.py --input Summer2026_primo_results_TIMESTAMP.xlsx
"""

import pandas as pd
import argparse
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── config ────────────────────────────────────────────────────────────────────

STAFF = ['AJ', 'AM', 'GM', 'NP', 'PL', 'RW']

PURCHASE_HEADERS = [
    'Course', 'Instructor', 'Title', 'Author', 'Edition',
    'Publication Year', 'Notes', 'Permanent Location', 'On Reserves?'
]

RESERVES_HEADERS = [
    'Title', 'Author', 'Call Number', 'Publication Year', 'ISBN',
    'Copies on Reserve', 'Instructor', 'Notes', 'Done?'
]

# Tab colours (openpyxl hex, no #)
TAB_COLOURS = {
    'Data':                   'D9E1F2',   # soft blue
    'AJ':                     'FCE4D6',   # peach
    'AM':                     'E2EFDA',   # light green
    'GM':                     'FFF2CC',   # yellow
    'NP':                     'F4CCFF',   # lavender
    'PL':                     'CCE5FF',   # sky blue
    'RW':                     'FFD9D9',   # pink
    'LW - Purchase':          'FFE0B2',   # orange-ish
    'TS - Put on Reserves':   'C8E6C9',   # green
    'TS - Take off Reserves': 'FFCCBC',   # coral
}

# Header fill colours (darker than tab)
HEADER_FILLS = {
    'Data':                   '9DC3E6',
    'AJ':                     'F4B183',
    'AM':                     'A9D18E',
    'GM':                     'FFD966',
    'NP':                     'D09FEB',
    'PL':                     '9DC3E6',
    'RW':                     'FF9999',
    'LW - Purchase':          'F4B183',
    'TS - Put on Reserves':   '70AD47',
    'TS - Take off Reserves': 'ED7D31',
}

# ── helpers ───────────────────────────────────────────────────────────────────

def thin_border():
    s = Side(style='thin', color='AAAAAA')
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, header_fill_hex):
    fill = PatternFill('solid', fgColor=header_fill_hex)
    font = Font(bold=True, color='FFFFFF' if _is_dark(header_fill_hex) else '000000')
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[1].height = 30

def _is_dark(hex_color):
    r, g, b = int(hex_color[0:2],16), int(hex_color[2:4],16), int(hex_color[4:6],16)
    return (0.299*r + 0.587*g + 0.114*b) < 128

def style_data_rows(ws, start_row=2):
    alt_fill = PatternFill('solid', fgColor='F2F2F2')
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(vertical='top', wrap_text=False)
            cell.border = thin_border()

def autofit_columns(ws, min_width=8, max_width=50):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))

def write_blank_sheet(ws, headers, header_fill_hex, tab_fill_hex):
    ws.append(headers)
    style_header_row(ws, header_fill_hex)
    autofit_columns(ws, min_width=12, max_width=40)
    ws.sheet_properties.tabColor = tab_fill_hex
    ws.freeze_panes = 'A2'

def write_data_sheet(ws, df, header_fill_hex, tab_fill_hex, sheet_label=''):
    # Write header + data
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))

    style_header_row(ws, header_fill_hex)
    style_data_rows(ws)
    autofit_columns(ws)
    ws.sheet_properties.tabColor = tab_fill_hex
    ws.freeze_panes = 'A2'

    if sheet_label:
        ws.sheet_view.showGridLines = True

# ── main ──────────────────────────────────────────────────────────────────────

def create_workbook(input_file: str, output_file: str):
    print(f"Reading: {input_file}")
    df = pd.read_excel(input_file)
    print(f"  {len(df)} rows loaded")

    # Remove lab/laboratory manuals — not collected
    manual_mask = df['Required_Title'].str.contains(
        r'lab\s*manual|laboratory\s*manual', case=False, na=False, regex=True
    )
    excluded = df[manual_mask]
    if len(excluded):
        print(f"  Excluded {len(excluded)} lab/laboratory manual(s):")
        for _, row in excluded.iterrows():
            print(f"    {row['Course']:10s}  {row['Required_Title']}")
    df = df[~manual_mask].reset_index(drop=True)
    print(f"  {len(df)} rows after exclusion")

    # Clean up column names for display
    df_display = df.copy()
    # Rename a few columns to be more human-friendly
    rename = {
        'Instructor_Name':  'Instructor',
        'Total_Enrollment': 'Enrollment',
        'ISBN_Original':    'ISBN',
        'ISBN_Clean':       'ISBN (cleaned)',
        'Required_Edition': 'Edition Required',
        'Required_Title':   'Title Required',
        'Found_Edition':    'Edition Found',
        'Publication_Year': 'Pub Year',
        'Edition_Match':    'Edition Match',
        'Has_Physical':     'Physical Copy?',
        'Has_Electronic':   'Electronic?',
        'Course_Assignment':'Course Assignment (Alma)',
    }
    df_display = df_display.rename(columns={k: v for k, v in rename.items() if k in df_display.columns})

    # Drop internal columns staff don't need
    drop_cols = ['EmplID', 'Class_Number', 'MMS_ID', 'Format', 'Availability', 'ISBN (cleaned)']
    df_display = df_display.drop(columns=[c for c in drop_cols if c in df_display.columns])

    # Sort by Recommendation priority then Course
    priority = {
        'Purchase Required': 1,
        'Wrong Edition - Verify or Purchase Needed': 2,
        'Check ISBN': 3,
        'May Need Purchase (verify manually)': 4,
        'Available - Verify Edition Manually': 5,
        'Already Available - Correct Edition': 6,
    }
    if 'Recommendation' in df_display.columns:
        df_display['_sort'] = df_display['Recommendation'].map(priority).fillna(9)
        df_display = df_display.sort_values(['_sort', 'Course']).drop(columns='_sort').reset_index(drop=True)

    # Divide rows among staff
    n = len(df_display)
    chunk = math.ceil(n / len(STAFF))
    staff_chunks = {
        s: df_display.iloc[i*chunk : (i+1)*chunk].copy()
        for i, s in enumerate(STAFF)
    }

    print(f"Divided {n} rows among {len(STAFF)} staff (~{chunk} each)")

    # ── build workbook ────────────────────────────────────────────────────────
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Data tab
        df_display.to_excel(writer, sheet_name='Data', index=False)

        # Staff tabs
        for s in STAFF:
            staff_chunks[s].to_excel(writer, sheet_name=s, index=False)

        # Blank sheets (just need the writer to create them)
        pd.DataFrame(columns=PURCHASE_HEADERS).to_excel(
            writer, sheet_name='LW - Purchase', index=False)
        pd.DataFrame(columns=RESERVES_HEADERS).to_excel(
            writer, sheet_name='TS - Put on Reserves', index=False)
        pd.DataFrame(columns=RESERVES_HEADERS).to_excel(
            writer, sheet_name='TS - Take off Reserves', index=False)

    # ── post-process styling with openpyxl ────────────────────────────────────
    wb = load_workbook(output_file)

    all_tabs = ['Data'] + STAFF + ['LW - Purchase', 'TS - Put on Reserves', 'TS - Take off Reserves']

    for tab in all_tabs:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        hfill = HEADER_FILLS.get(tab, 'AAAAAA')
        tfill = TAB_COLOURS.get(tab, 'FFFFFF')

        style_header_row(ws, hfill)
        ws.sheet_properties.tabColor = tfill
        ws.freeze_panes = 'A2'

        if tab not in ('LW - Purchase', 'TS - Put on Reserves', 'TS - Take off Reserves'):
            style_data_rows(ws)

        autofit_columns(ws)

    wb.save(output_file)
    print(f"Saved: {output_file}")

    # Summary
    print()
    print("Tabs created:")
    for tab in all_tabs:
        if tab in ('LW - Purchase', 'TS - Put on Reserves', 'TS - Take off Reserves'):
            print(f"  {tab:30s}  (blank template)")
        elif tab == 'Data':
            print(f"  {tab:30s}  {len(df_display)} rows (full results)")
        else:
            rows = len(staff_chunks.get(tab, []))
            print(f"  {tab:30s}  {rows} rows")


def main():
    parser = argparse.ArgumentParser(description='Create staff review workbook')
    parser.add_argument('--input',  required=True, help='Primo results Excel file')
    parser.add_argument('--output', default=None,  help='Output file path')
    parser.add_argument('--staff',  default=None,
                        help='Comma-separated staff initials (default: AJ,AM,GM,NP,PL,RW)')
    args = parser.parse_args()

    if args.staff:
        global STAFF
        STAFF = [s.strip() for s in args.staff.split(',')]

    if not args.output:
        p = Path(args.input)
        args.output = str(p.parent / p.name.replace('primo_results', 'staff_workbook'))

    create_workbook(args.input, args.output)


if __name__ == '__main__':
    main()
