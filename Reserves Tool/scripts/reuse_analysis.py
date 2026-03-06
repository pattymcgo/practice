"""
Reserves Reuse Analysis - CUNY BMCC
-------------------------------------
Analyzes which titles are reused across multiple semesters using the
Alma citations export (Course Terms field).

Reuse tiers:
  - Long-Term Staple:    used in 5 or more semesters
  - Frequently Reused:   used in 3-4 semesters
  - Occasional:          used in 2 semesters
  - Single Semester:     used in 1 semester only

Output: Excel report with reuse tiers and per-course breakdown.

Usage:
    python scripts/reuse_analysis.py

Author: Patty (with Claude Code assistance)
Date: 2026-03-05
"""

import pandas as pd
import re
from datetime import datetime
import os
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ============================================================================
# FILE PATHS
# ============================================================================

BASE_DIR       = "/Users/patty_home/Desktop/Agentic AI/Reserves Tool"
CITATIONS_FILE = os.path.join(BASE_DIR, "data/SP26 Citations.xlsx")
REPORTS_DIR    = os.path.join(BASE_DIR, "reports")

# Chronological order for sorting terms consistently
TERM_ORDER = {
    'Summer 2023':  1,
    'Fall 2023':    2,
    'Winter 2024':  3,
    'Spring 2024':  4,
    'Summer 2024':  5,
    'Fall 2024':    6,
    'Winter 2025':  7,
    'Spring 2025':  8,
    'Summer 2025':  9,
    'Fall 2025':   10,
    'Winter 2026': 11,
    'Spring 2026': 12,
}

# Reuse tier thresholds
STAPLE_MIN     = 5   # 5+ semesters = Long-Term Staple
FREQUENT_MIN   = 3   # 3-4 semesters = Frequently Reused
OCCASIONAL_MIN = 2   # 2 semesters   = Occasional
# 1 semester = Single Semester


# ============================================================================
# HELPERS
# ============================================================================

def normalize_title(title):
    """Lowercase, strip articles and punctuation for grouping duplicate titles."""
    if pd.isna(title):
        return ""
    t = str(title).lower()
    t = re.sub(r'^(the|a|an)\s+', '', t)
    t = re.sub(r'[^a-z0-9\s]', '', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def parse_terms(terms_str):
    """Parse a comma-separated Course Terms string into a sorted list."""
    if pd.isna(terms_str):
        return []
    terms = [t.strip() for t in str(terms_str).split(',') if t.strip()]
    # Sort chronologically using TERM_ORDER; unknown terms go last
    terms = sorted(set(terms), key=lambda t: TERM_ORDER.get(t, 99))
    return terms


def assign_tier(semester_count):
    """Return the reuse tier label for a given semester count."""
    if semester_count >= STAPLE_MIN:
        return 'Long-Term Staple (5+ semesters)'
    elif semester_count >= FREQUENT_MIN:
        return 'Frequently Reused (3-4 semesters)'
    elif semester_count >= OCCASIONAL_MIN:
        return 'Occasional (2 semesters)'
    else:
        return 'Single Semester'


def clean_course_code(code):
    """Strip instructor name from course code, e.g. 'ENG 201 (Noveno)' -> 'ENG 201'."""
    if pd.isna(code):
        return ""
    return re.sub(r'\s*\(.*?\)', '', str(code)).strip()


# ============================================================================
# MAIN ANALYSIS
# ============================================================================

def build_title_reuse(df):
    """
    Group by normalized title to find reuse across all courses and semesters.
    Each row in the output = one unique title, with all terms and courses combined.
    """
    rows = []
    for _, row in df.iterrows():
        terms = parse_terms(row.get('Course Terms', ''))
        rows.append({
            'Title':            row.get('Citation Title', ''),
            'Title_Norm':       normalize_title(row.get('Citation Title', '')),
            'ISBN':             row.get('ISBN', ''),
            'ISBN_13':          row.get('ISBN (13)', ''),
            'Course_Code':      clean_course_code(row.get('Course Code', '')),
            'Course_Name':      row.get('Course Name', ''),
            'Academic_Dept':    row.get('Academic Department', ''),
            'Terms_List':       terms,
            'Semester_Count':   len(terms),
            'Is_SP26':          'Spring 2026' in terms,
        })

    cit_df = pd.DataFrame(rows)

    # --- Title-level aggregation ---
    def join_unique_sorted(series):
        vals = sorted({str(v).strip() for v in series.dropna() if str(v).strip()})
        return ', '.join(vals)

    def union_terms(series):
        all_terms = set()
        for lst in series:
            all_terms.update(lst)
        return sorted(all_terms, key=lambda t: TERM_ORDER.get(t, 99))

    title_groups = cit_df.groupby('Title_Norm', dropna=False).agg(
        Title            =('Title',         'first'),
        ISBN             =('ISBN',          'first'),
        ISBN_13          =('ISBN_13',       'first'),
        Courses          =('Course_Code',   join_unique_sorted),
        Departments      =('Academic_Dept', join_unique_sorted),
        All_Terms_List   =('Terms_List',    union_terms),
        Active_SP26      =('Is_SP26',       'any'),
        Num_Course_Sections=('Course_Code', 'count'),
    ).reset_index(drop=True)

    title_groups['Semester_Count'] = title_groups['All_Terms_List'].apply(len)
    title_groups['Semesters_Used'] = title_groups['All_Terms_List'].apply(
        lambda lst: ', '.join(lst)
    )
    title_groups['First_Seen']  = title_groups['All_Terms_List'].apply(
        lambda lst: lst[0] if lst else ''
    )
    title_groups['Last_Seen']   = title_groups['All_Terms_List'].apply(
        lambda lst: lst[-1] if lst else ''
    )
    title_groups['Reuse_Tier']  = title_groups['Semester_Count'].apply(assign_tier)
    title_groups['On_SP26']     = title_groups['Active_SP26'].apply(
        lambda x: 'Yes' if x else 'No'
    )

    # Final column order
    out_cols = [
        'Reuse_Tier', 'Semester_Count', 'Title', 'ISBN', 'ISBN_13',
        'Courses', 'Departments', 'Num_Course_Sections',
        'Semesters_Used', 'First_Seen', 'Last_Seen', 'On_SP26',
    ]
    return title_groups[out_cols].sort_values(
        ['Semester_Count', 'Title'], ascending=[False, True]
    ).reset_index(drop=True)


def build_course_title_reuse(df):
    """
    Per-course per-title reuse: how many semesters has THIS COURSE used THIS BOOK?
    Useful for identifying course-specific reading list staples.
    """
    rows = []
    for _, row in df.iterrows():
        terms = parse_terms(row.get('Course Terms', ''))
        rows.append({
            'Course_Code':   clean_course_code(row.get('Course Code', '')),
            'Course_Name':   row.get('Course Name', ''),
            'Department':    row.get('Academic Department', ''),
            'Title':         row.get('Citation Title', ''),
            'Title_Norm':    normalize_title(row.get('Citation Title', '')),
            'ISBN':          row.get('ISBN', ''),
            'Terms_List':    terms,
            'Is_SP26':       'Spring 2026' in terms,
        })

    cit_df = pd.DataFrame(rows)

    def union_terms(series):
        all_terms = set()
        for lst in series:
            all_terms.update(lst)
        return sorted(all_terms, key=lambda t: TERM_ORDER.get(t, 99))

    groups = cit_df.groupby(['Course_Code', 'Title_Norm'], dropna=False).agg(
        Course_Name    =('Course_Name',  'first'),
        Department     =('Department',   'first'),
        Title          =('Title',        'first'),
        ISBN           =('ISBN',         'first'),
        All_Terms_List =('Terms_List',   union_terms),
        Active_SP26    =('Is_SP26',      'any'),
    ).reset_index()

    groups['Semester_Count'] = groups['All_Terms_List'].apply(len)
    groups['Semesters_Used'] = groups['All_Terms_List'].apply(lambda lst: ', '.join(lst))
    groups['Reuse_Tier']     = groups['Semester_Count'].apply(assign_tier)
    groups['On_SP26']        = groups['Active_SP26'].apply(lambda x: 'Yes' if x else 'No')

    out_cols = [
        'Reuse_Tier', 'Semester_Count', 'Course_Code', 'Course_Name',
        'Department', 'Title', 'ISBN', 'Semesters_Used', 'On_SP26',
    ]
    # Only show titles used 2+ semesters for this view (single-use not interesting here)
    result = groups[groups['Semester_Count'] >= 2][out_cols].sort_values(
        ['Semester_Count', 'Course_Code', 'Title'], ascending=[False, True, True]
    ).reset_index(drop=True)
    return result


# ============================================================================
# EXCEL STYLING
# ============================================================================

TAB_COLORS = {
    'Summary':             ('1F4E79', 'DCE6F1'),  # Dark blue / light blue
    'Long_Term_Staples':   ('375623', 'E2EFDA'),  # Dark green / light green
    'Frequently_Reused':   ('7F6000', 'FFEB9C'),  # Dark gold / light yellow
    'Occasional':          ('833C00', 'FCE4D6'),  # Dark brown / light peach
    'All_Titles':          ('1F4E79', 'DCE6F1'),  # Dark blue / light blue
    'By_Course_Title':     ('17375E', 'DDEBF7'),  # Dark teal / light blue
}

TIER_FILLS = {
    'Long-Term Staple (5+ semesters)':    PatternFill('solid', fgColor='E2EFDA'),
    'Frequently Reused (3-4 semesters)':  PatternFill('solid', fgColor='FFEB9C'),
    'Occasional (2 semesters)':           PatternFill('solid', fgColor='FCE4D6'),
    'Single Semester':                    PatternFill('solid', fgColor='FFFFFF'),
}


def style_sheet(ws, tab_name):
    """Apply color formatting: colored header, zebra striping, freeze, auto-width."""
    header_hex, zebra_hex = TAB_COLORS.get(tab_name, ('1F4E79', 'DCE6F1'))
    header_fill = PatternFill('solid', fgColor=header_hex)
    zebra_fill  = PatternFill('solid', fgColor=zebra_hex)
    white_fill  = PatternFill('solid', fgColor='FFFFFF')
    header_font = Font(bold=True, color='FFFFFF')
    section_font = Font(bold=True, color=header_hex)

    # Find Reuse_Tier column index if present
    tier_col_idx = None
    header_row = [cell.value for cell in ws[1]]
    if 'Reuse_Tier' in header_row:
        tier_col_idx = header_row.index('Reuse_Tier') + 1

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical='top')

        if row_idx == 1:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
        else:
            first_val = str(row[0].value or '')
            if tab_name == 'Summary' and first_val.startswith('---'):
                for cell in row:
                    cell.font = section_font
                    cell.fill = zebra_fill
            else:
                fill = zebra_fill if row_idx % 2 == 0 else white_fill
                for cell in row:
                    cell.fill = fill

            # Color entire row by tier if Reuse_Tier column present
            if tier_col_idx:
                tier_cell = row[tier_col_idx - 1]
                tier_val = str(tier_cell.value or '')
                if tier_val in TIER_FILLS:
                    for cell in row:
                        cell.fill = TIER_FILLS[tier_val]

    ws.freeze_panes = 'A2'

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 70)


# ============================================================================
# MAIN
# ============================================================================

def main():
    timestamp   = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(REPORTS_DIR, f"reuse_analysis_{timestamp}.xlsx")

    print("=" * 65)
    print("RESERVES REUSE ANALYSIS - CUNY BMCC")
    print("=" * 65)
    print()

    print(f"Loading citations file: {CITATIONS_FILE}")
    df = pd.read_excel(CITATIONS_FILE)
    print(f"  Loaded {len(df)} citations")
    print()

    print("Building title-level reuse table...")
    title_reuse = build_title_reuse(df)
    print(f"  Unique titles found: {len(title_reuse)}")

    staples   = title_reuse[title_reuse['Reuse_Tier'].str.startswith('Long-Term')]
    frequent  = title_reuse[title_reuse['Reuse_Tier'].str.startswith('Frequently')]
    occasional = title_reuse[title_reuse['Reuse_Tier'].str.startswith('Occasional')]
    single    = title_reuse[title_reuse['Reuse_Tier'].str.startswith('Single')]

    print(f"    Long-Term Staples (5+ semesters): {len(staples)}")
    print(f"    Frequently Reused (3-4 semesters): {len(frequent)}")
    print(f"    Occasional (2 semesters):           {len(occasional)}")
    print(f"    Single Semester:                    {len(single)}")
    print()

    print("Building per-course title reuse table...")
    course_reuse = build_course_title_reuse(df)
    print(f"  Course+title combinations used 2+ semesters: {len(course_reuse)}")
    print()

    # -------------------------------------------------------------------------
    # Summary tab
    # -------------------------------------------------------------------------
    sp26_staples  = staples[staples['On_SP26'] == 'Yes']
    sp26_frequent = frequent[frequent['On_SP26'] == 'Yes']

    summary_data = {
        'Category': [
            '--- TITLE REUSE OVERVIEW ---',
            'Total unique titles in citations file',
            'Long-Term Staples (used 5+ semesters)',
            '  - Still active in Spring 2026',
            '  - No longer active (consider keeping for future)',
            'Frequently Reused (used 3-4 semesters)',
            '  - Still active in Spring 2026',
            'Occasional (used 2 semesters)',
            'Single Semester only',
            '',
            '--- SEMESTERS IN THE DATA ---',
            'Earliest term seen',
            'Latest term seen',
            'Total distinct terms',
            '',
            '--- PER-COURSE REUSE ---',
            'Course+title pairs reused 2+ semesters',
            'Course+title pairs that are Long-Term Staples',
        ],
        'Count': [
            '',
            len(title_reuse),
            len(staples),
            len(sp26_staples),
            len(staples) - len(sp26_staples),
            len(frequent),
            len(sp26_frequent),
            len(occasional),
            len(single),
            '',
            '',
            min(TERM_ORDER, key=lambda t: TERM_ORDER[t]),
            max(TERM_ORDER, key=lambda t: TERM_ORDER[t]),
            len(TERM_ORDER),
            '',
            '',
            len(course_reuse),
            len(course_reuse[course_reuse['Reuse_Tier'].str.startswith('Long-Term')]),
        ],
        'Notes': [
            '',
            'Grouped by normalized title (articles/punctuation removed)',
            'Strong candidates to keep on permanent reserve',
            'Currently on Spring 2026 reading lists',
            'Not on SP26 — but historically high-use; review before removing',
            'Used consistently but not yet at staple level',
            'Currently on Spring 2026 reading lists',
            'Used in exactly 2 semesters',
            'Only appeared once — standard review applies',
            '',
            '',
            '',
            '',
            'Summer 2023 through Spring 2026',
            '',
            '',
            'Title reused by same course across multiple semesters',
            'Same course has used same title 5+ semesters',
        ],
    }
    df_summary = pd.DataFrame(summary_data)

    # -------------------------------------------------------------------------
    # Write Excel
    # -------------------------------------------------------------------------
    print(f"Writing report to: {output_file}")
    os.makedirs(REPORTS_DIR, exist_ok=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        staples.to_excel(writer, sheet_name='Long_Term_Staples', index=False)
        frequent.to_excel(writer, sheet_name='Frequently_Reused', index=False)
        occasional.to_excel(writer, sheet_name='Occasional', index=False)
        title_reuse.to_excel(writer, sheet_name='All_Titles', index=False)
        course_reuse.to_excel(writer, sheet_name='By_Course_Title', index=False)

        for tab_name, ws in writer.sheets.items():
            style_sheet(ws, tab_name)

    print()
    print("=" * 65)
    print("DONE")
    print("=" * 65)
    print(f"Report saved to: {output_file}")
    print()
    print("Tabs in the report:")
    print("  Summary           - Overview counts and tier breakdown")
    print("  Long_Term_Staples - Titles used 5+ semesters (green)")
    print("  Frequently_Reused - Titles used 3-4 semesters (yellow)")
    print("  Occasional        - Titles used in exactly 2 semesters (peach)")
    print("  All_Titles        - Every title sorted by reuse count")
    print("  By_Course_Title   - Per-course reuse (same course + same book, 2+ semesters)")


if __name__ == "__main__":
    main()
