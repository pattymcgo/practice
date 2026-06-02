"""
Course Reserves Comparison Tool - CUNY BMCC
--------------------------------------------
Compares the current Alma course reserves (SP26 Citations export) against
the Spring 2026 merged course/textbook dataset to identify:

  1. Items confirmed correct on reserves for Spring 2026
  2. Items missing from reserves (in merged dataset but not in citations)
  3. Items to potentially remove (on reading lists but not in Spring 2026)
     - Priority: items physically on the RESE shelf
  4. Items needing review (course/edition mismatches)
  5. Course roster check (via Alma Courses API):
     - Active Spring 2026 Alma courses with no textbook request submitted
     - Textbook requests for courses with no active Alma course record

OUTPUT: Excel report with tabs. NO changes are made to Alma.

Usage:
    python scripts/compare_reserves.py

Author: Patty (with Claude Code assistance)
"""

import json
import time
import pandas as pd
import re
import requests
from datetime import datetime
import os
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ============================================================================
# FILE PATHS
# ============================================================================

BASE_DIR        = "/Users/patty_home/Desktop/Agentic AI/Reserves Tool"
CITATIONS_FILE  = os.path.join(BASE_DIR, "data/SP26 Citations.xlsx")
MERGED_FILE     = os.path.join(BASE_DIR, "data/merged_course_textbooks_CLEANED.xlsx")
REPORTS_DIR     = os.path.join(BASE_DIR, "reports")
CONFIG_PATH     = os.path.join(BASE_DIR, "scripts/isbn_search/config.json")

ALMA_BASE = "https://api-na.hosted.exlibrisgroup.com"

# Labels Alma might use for Spring 2026 in the term field
SPRING_2026_ALIASES = {"spring 2026", "2026 spring", "sp26", "s26", "spring26", "spring 26"}

# Chronological ranking of academic terms (higher number = more recent)
# Used for sorting and determining the removal cutoff
TERM_ORDER = {
    'Summer 2023':  1,
    'Fall 2023':    2,
    'Winter 2024':  3,
    'Spring 2024':  4,
    'Summer 2024':  5,
    'Fall 2024':    6,   # <-- removal cutoff: only items at or below this rank
    'Winter 2025':  7,
    'Spring 2025':  8,
    'Summer 2025':  9,
    'Fall 2025':   10,
    'Winter 2026': 11,
    'Spring 2026': 12,
}

# Items last used in this term or earlier are candidates for removal
# (approx. 2 years back from Spring 2026 = Spring 2024 and older)
REMOVAL_CUTOFF_TERM = 'Spring 2024'
REMOVAL_CUTOFF_RANK = TERM_ORDER[REMOVAL_CUTOFF_TERM]

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def clean_course_code(code):
    """
    Strip instructor name from course code.
    e.g. 'ENG 201 (Noveno)' -> 'ENG 201'
    """
    if pd.isna(code):
        return ""
    return re.sub(r'\s*\(.*?\)', '', str(code)).strip()


def normalize_title(title):
    """
    Lowercase, remove articles and punctuation for fuzzy title matching.
    e.g. 'The Art of War' -> 'art war'
    """
    if pd.isna(title):
        return ""
    t = str(title).lower()
    # Remove leading articles
    t = re.sub(r'^(the|a|an)\s+', '', t)
    # Remove all non-alphanumeric characters
    t = re.sub(r'[^a-z0-9\s]', '', t)
    # Collapse whitespace
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def extract_isbns(isbn_field):
    """
    Parse a semicolon-separated ISBN field into a set of cleaned ISBNs.
    Handles both ISBN-10 and ISBN-13.
    """
    if pd.isna(isbn_field):
        return set()
    isbns = set()
    for raw in str(isbn_field).split(';'):
        cleaned = re.sub(r'[\s\-]', '', raw.strip())
        # Keep only digits and trailing X (ISBN-10)
        cleaned = re.sub(r'[^0-9Xx]', '', cleaned)
        if len(cleaned) in (10, 13):
            isbns.add(cleaned.upper())
    return isbns


def isbns_overlap(set_a, set_b):
    """Return True if any ISBN in set_a appears in set_b."""
    return bool(set_a & set_b)


def fix_id_columns(df):
    """
    Convert large integer ID columns to strings so Excel doesn't display
    them in scientific notation (e.g. 3.147e+13 → '31470000000000').
    Works on a copy of the dataframe.
    """
    df = df.copy()
    for col in ['MMS ID', 'Citation ID', 'Holdings ID', 'Item ID', 'Barcode']:
        if col in df.columns:
            df[col] = df[col].apply(
                lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ('', 'nan') else ''
            )
    return df


def parse_edition_from_imprint(imprint):
    """
    Try to extract a publication year from the Type/Creator/Imprint field.
    e.g. 'Book (Bedford/St Martin s [2020])' -> '2020'
    This is used as a rough proxy for edition comparison.
    """
    if pd.isna(imprint):
        return ""
    years = re.findall(r'\b(19|20)\d{2}\b', str(imprint))
    return years[-1] if years else ""


# ============================================================================
# ALMA COURSES API  (course roster check)
# ============================================================================

def load_alma_api_key():
    """Load the Alma API key from config.json. Returns the key or None."""
    try:
        with open(CONFIG_PATH) as f:
            return json.load(f).get("alma_api_key")
    except FileNotFoundError:
        return None


def fetch_active_sp26_courses(api_key):
    """
    Fetch all courses from Alma and return only those that are:
      - Status = ACTIVE
      - Belong to Spring 2026 (term label or date window)

    Returns a list of course dicts.
    """
    url     = f"{ALMA_BASE}/almaws/v1/courses"
    headers = {"Authorization": f"apikey {api_key}", "Accept": "application/json"}
    offset  = 0
    limit   = 100
    total   = None
    all_courses = []

    print("  Fetching all courses from Alma Courses API...")

    while True:
        resp = requests.get(
            url, headers=headers,
            params={"limit": limit, "offset": offset},
            timeout=30,
        )
        if resp.status_code != 200:
            print(f"  WARNING: Alma API returned {resp.status_code}. Skipping course roster check.")
            return []

        data = resp.json()
        if total is None:
            total = int(data.get("total_record_count", 0))

        courses = data.get("course", [])
        if not isinstance(courses, list):
            courses = [courses] if courses else []

        all_courses.extend(courses)
        print(f"    Fetched {len(all_courses):,} / {total:,} ...", end="\r")

        if len(all_courses) >= total or not courses:
            break

        offset += limit
        time.sleep(0.2)

    print(f"\n  Total courses fetched: {len(all_courses):,}")

    # Filter to active Spring 2026 courses
    sp26 = []
    for c in all_courses:
        # Status check
        status = c.get("status", "")
        if isinstance(status, dict):
            status = status.get("value", "")
        if str(status).upper() != "ACTIVE":
            continue

        # Spring 2026 term check
        terms = c.get("term", [])
        if not isinstance(terms, list):
            terms = [terms] if terms else []

        is_sp26 = False
        for t in terms:
            val = (t.get("value") or t.get("desc") or "").lower().strip()
            if any(alias in val for alias in SPRING_2026_ALIASES):
                is_sp26 = True
                break

        # Fallback: check if start_date OR end_date falls in Spring 2026 window (Jan–Jun 2026)
        # NOTE: Must check both fields because many courses span multiple semesters and
        # have a start_date from years ago — the end_date is what places them in SP26.
        if not is_sp26:
            for date_field in ("start_date", "end_date"):
                raw = c.get(date_field, "")
                if not raw:
                    continue
                try:
                    d = datetime.fromisoformat(raw.rstrip("Z").split("T")[0])
                    if datetime(2026, 1, 1) <= d <= datetime(2026, 6, 30):
                        is_sp26 = True
                        break
                except ValueError:
                    pass

        if is_sp26:
            sp26.append(c)

    print(f"  Active Spring 2026 courses in Alma: {len(sp26):,}")
    return sp26


def build_course_roster_check(merged_df, alma_courses):
    """
    Cross-reference active Spring 2026 Alma courses against the textbook
    spreadsheet (merged_df).

    Returns two DataFrames:
      alma_no_request  -- Alma courses with NO textbook request in merged_df
      roster_no_alma   -- Merged courses with NO matching active Alma course
    """
    # Build set of course codes from the textbook spreadsheet
    roster_codes = set(merged_df['Course_Clean'].str.strip().str.upper().dropna())

    # Build set of course codes from Alma (strip instructor names)
    alma_code_map = {}  # clean_code -> list of course dicts
    for c in alma_courses:
        raw_code   = c.get("code", "")
        clean_code = clean_course_code(raw_code).upper()
        alma_code_map.setdefault(clean_code, []).append(c)

    alma_codes = set(alma_code_map.keys())

    # --- Alma courses with no textbook request ---
    alma_no_request_rows = []
    for clean_code, courses in sorted(alma_code_map.items()):
        if clean_code in roster_codes:
            continue
        for c in courses:
            # Get term label
            terms = c.get("term", [])
            if isinstance(terms, list):
                term_labels = [t.get("desc") or t.get("value") or "" for t in terms]
                term_str = " | ".join(t for t in term_labels if t)
            else:
                term_str = str(terms)

            # Get instructor
            instructors = c.get("instructor", [])
            if not isinstance(instructors, list):
                instructors = [instructors] if instructors else []
            instr_str = "; ".join(
                f"{i.get('last_name','')}, {i.get('first_name','')}".strip(", ")
                for i in instructors
            )

            alma_no_request_rows.append({
                "Course_Code":  clean_code,
                "Course_Name":  c.get("name", ""),
                "Section":      c.get("section", ""),
                "Term":         term_str,
                "Instructor":   instr_str,
                "Enrollment":   c.get("participants_number", ""),
                "Start_Date":   c.get("start_date", ""),
                "End_Date":     c.get("end_date", ""),
                "Flag":         "Active in Alma — no textbook request found",
            })

    # --- Textbook requests with no matching Alma course ---
    roster_no_alma_rows = []
    courses_missing = sorted(roster_codes - alma_codes)

    for course_code in courses_missing:
        subset = merged_df[merged_df['Course_Clean'].str.upper() == course_code]
        sections    = ", ".join(sorted(subset['Section'].astype(str).unique()))
        instructors = "; ".join(sorted(subset['Instructor_Name'].dropna().unique()))
        enrollment  = subset['Total_Enrollment'].apply(
            lambda x: pd.to_numeric(x, errors='coerce')
        ).sum()
        titles = "; ".join(subset['Title'].dropna().unique()[:3])  # first 3 titles
        if len(subset['Title'].dropna().unique()) > 3:
            titles += " ..."

        roster_no_alma_rows.append({
            "Course_Code":       course_code,
            "Sections":          sections,
            "Instructors":       instructors,
            "Total_Enrollment":  int(enrollment) if enrollment > 0 else "",
            "Textbook_Count":    len(subset),
            "Titles_Preview":    titles,
            "Flag":              "In textbook spreadsheet — no active Alma course found",
        })

    return pd.DataFrame(alma_no_request_rows), pd.DataFrame(roster_no_alma_rows)


def build_section_instructor_check(merged_df, alma_courses):
    """
    For courses that exist in BOTH the textbook spreadsheet AND Alma,
    check whether the sections and instructors match.

    Section logic:
      - If Alma's section field is blank, that is flagged as needing manual
        verification — it may mean all sections or it may be missing data.
      - If Alma lists specific sections, they are compared to the spreadsheet.

    Instructor logic:
      - Names are normalized to uppercase last name for comparison (handles
        format differences like 'Kleinman,Harry' vs 'KLEINMAN, HARRY').
      - Empty Alma instructor entries (', ') are ignored.

    Returns a DataFrame of courses with at least one mismatch or gap.
    """
    roster_codes = set(merged_df['Course_Clean'].str.strip().str.upper().dropna())

    alma_code_map = {}
    for c in alma_courses:
        clean_code = clean_course_code(c.get("code", "")).upper()
        alma_code_map.setdefault(clean_code, []).append(c)

    matched_codes = roster_codes & set(alma_code_map.keys())

    rows = []
    for course_code in sorted(matched_codes):
        subset = merged_df[merged_df['Course_Clean'].str.upper() == course_code]

        # --- Textbook sections and instructors ---
        roster_sections = {
            s.strip() for s in subset['Section'].astype(str).unique()
            if s.strip() and s.strip().lower() != 'nan'
        }
        roster_instr_raw = {
            str(n).strip() for n in subset['Instructor_Name'].dropna().unique()
            if str(n).strip()
        }
        roster_lastnames = {
            n.split(",")[0].strip().upper() for n in roster_instr_raw if n
        }

        # --- Alma sections and instructors ---
        alma_section_set  = set()
        alma_instr_raw    = set()
        alma_section_blank = False  # True if any Alma record has no section value

        for c in alma_code_map[course_code]:
            sec = str(c.get("section", "") or "").strip()
            if not sec:
                alma_section_blank = True
            else:
                for s in sec.split(","):
                    s = s.strip()
                    if s:
                        alma_section_set.add(s)

            instructors = c.get("instructor", [])
            if not isinstance(instructors, list):
                instructors = [instructors] if instructors else []
            for i in instructors:
                last  = (i.get("last_name",  "") or "").strip()
                first = (i.get("first_name", "") or "").strip()
                if last:  # skip blank entries like ', '
                    alma_instr_raw.add(f"{last}, {first}".strip(", ").upper())

        alma_lastnames = {
            n.split(",")[0].strip().upper() for n in alma_instr_raw if n
        }

        # --- Section comparison ---
        if alma_section_blank and not alma_section_set:
            # No section data at all in Alma — flag for manual check
            sections_missing_from_alma   = set()
            sections_missing_from_roster = set()
            section_status = "No section data in Alma — verify manually"
        elif alma_section_blank:
            # Some records have sections, at least one is blank — compare what we have
            sections_missing_from_alma   = roster_sections - alma_section_set
            sections_missing_from_roster = alma_section_set - roster_sections
            section_status = "Partial — at least one Alma record has no section"
        else:
            sections_missing_from_alma   = roster_sections - alma_section_set
            sections_missing_from_roster = alma_section_set - roster_sections
            section_status = "OK" if not sections_missing_from_alma and not sections_missing_from_roster else "Mismatch"

        # --- Instructor comparison (last name only) ---
        instr_missing_from_alma   = roster_lastnames - alma_lastnames
        instr_missing_from_roster = alma_lastnames - roster_lastnames
        instr_status = "OK" if not instr_missing_from_alma and not instr_missing_from_roster else "Mismatch"

        has_issue = section_status != "OK" or instr_status != "OK"
        if has_issue:
            rows.append({
                "Course_Code":                    course_code,
                "Section_Status":                 section_status,
                "Roster_Sections":                ", ".join(sorted(roster_sections)),
                "Alma_Sections":                  ", ".join(sorted(alma_section_set)) if alma_section_set else "(blank)",
                "Sections_In_Roster_Not_Alma":    ", ".join(sorted(sections_missing_from_alma)),
                "Sections_In_Alma_Not_Roster":    ", ".join(sorted(sections_missing_from_roster)),
                "Instructor_Status":              instr_status,
                "Roster_Instructors":             "; ".join(sorted(roster_instr_raw)),
                "Alma_Instructors":               "; ".join(sorted(alma_instr_raw)),
                "Instructors_In_Roster_Not_Alma": "; ".join(sorted(instr_missing_from_alma)),
                "Instructors_In_Alma_Not_Roster": "; ".join(sorted(instr_missing_from_roster)),
            })

    return pd.DataFrame(rows)


# ============================================================================
# LOAD DATA
# ============================================================================

def load_data():
    print("Loading files...")
    citations = pd.read_excel(CITATIONS_FILE)
    merged    = pd.read_excel(MERGED_FILE)
    print(f"  SP26 Citations: {len(citations)} rows")
    print(f"  Merged dataset: {len(merged)} rows")
    return citations, merged


# ============================================================================
# PREPARE CITATIONS
# ============================================================================

def prepare_citations(citations):
    """
    Add helper columns to the citations DataFrame.
    """
    df = citations.copy()

    # Clean course code (strip instructor name)
    df['Course_Clean'] = df['Course Code'].apply(clean_course_code)

    # Flag: is this citation active for Spring 2026?
    df['Is_SP26'] = df['Course Terms'].str.contains('Spring 2026', na=False)

    # Flag: is the item on the temporary RESE reserve shelf?
    # (RESE is a temporary location - books go on/off reserves from here)
    df['On_RESE_Shelf'] = df['Availability'].str.contains('RESE', na=False)

    # Flag: is the item in the permanent CLOSED location?
    # (CLOSED is permanent - book stays there regardless of reserve status)
    df['In_CLOSED_Location'] = df['Availability'].str.contains('CLOSED', na=False)

    # Combine both ISBN columns into one set per row
    df['ISBN_Set'] = df.apply(
        lambda r: extract_isbns(r['ISBN']) | extract_isbns(r['ISBN (13)']),
        axis=1
    )

    # Normalized title for fuzzy matching
    df['Title_Norm'] = df['Citation Title'].apply(normalize_title)

    # Parse year from imprint as proxy for edition
    df['Pub_Year'] = df['Type / Creator / Imprint'].apply(parse_edition_from_imprint)

    # Most recent term on this citation (uses global TERM_ORDER for correct chronology)
    def most_recent_term(terms_str):
        if pd.isna(terms_str):
            return 'Unknown'
        terms = [t.strip() for t in str(terms_str).split(',')]
        ranked = [(TERM_ORDER.get(t, 0), t) for t in terms]
        return max(ranked, key=lambda x: x[0])[1]

    df['Most_Recent_Term'] = df['Course Terms'].apply(most_recent_term)

    return df


# ============================================================================
# TITLE FILTERS
# ============================================================================

# Titles the library does not purchase - rows with these terms are excluded
EXCLUDE_TITLE_TERMS = [
    'lab manual',
    'lab. manual',
    'laboratory manual',
    'laboratory',
    'with connect',
    'access with connect',
    'connect online access',
    'conect core',
    'w/connect',
    '+online access',
    '+ online access',
    'with quickbooks',
]

# Publishers the library does not purchase from - rows with these publishers are excluded
EXCLUDE_PUBLISHER_TERMS = [
    'openstax',
    'open stax',
    'opentextbook',
]

# Titles indicating open educational resources - reclassified from Book to OER
OER_TITLE_TERMS = ['oer', 'openstax', 'open stax']


# ============================================================================
# PRIMO CATALOG LOOKUP
# ============================================================================

PRIMO_HOST  = "https://cuny-bm.primo.exlibrisgroup.com"
PRIMO_VID   = "01CUNY_BM:CUNY_BM"
PRIMO_SCOPE = "MyInst_and_CI"


def search_primo(isbn):
    """
    Search CUNY BMCC Primo catalog for a single ISBN.
    Returns a dict with availability info, or None if not found.
    """
    try:
        response = requests.get(
            f"{PRIMO_HOST}/primaws/rest/pub/pnxs",
            params={'q': f'isbn,exact,{isbn}', 'vid': PRIMO_VID,
                    'scope': PRIMO_SCOPE, 'lang': 'en'},
            timeout=15
        )
        if response.status_code != 200:
            return None
        data = response.json()
        if data.get('info', {}).get('total', 0) == 0:
            return None

        doc     = data['docs'][0]
        display = doc.get('pnx', {}).get('display', {})
        avail   = ' '.join(display.get('availlibrary', [])).lower()

        return {
            'Catalog_Title':    (display.get('title',   [''])[0]),
            'Has_Physical':     'main campus' in avail or ('available' in avail and 'electronic' not in avail),
            'Has_Electronic':   'electronic' in avail or 'online' in avail,
            'Catalog_Format':   (display.get('type',    [''])[0]),
            'Pub_Year':         (display.get('creationdate', [''])[0]),
        }
    except Exception:
        return None


def lookup_missing_books_in_primo(df):
    """
    For each unique ISBN in the Missing_Books dataframe, search Primo.
    Adds columns: In_Catalog, Catalog_Title, Has_Physical, Has_Electronic,
                  Catalog_Format, Catalog_Pub_Year, Catalog_Action.
    Deduplicates ISBN lookups so each ISBN is only searched once.
    """
    unique_isbns = df['ISBN'].dropna().unique()
    total = len(unique_isbns)
    print(f"  Searching Primo for {total} unique ISBNs in Missing_Books...")

    isbn_results = {}
    for i, isbn in enumerate(unique_isbns, 1):
        if i % 25 == 0:
            print(f"    Progress: {i}/{total}")
        isbn_clean = re.sub(r'[\s\-]', '', str(isbn))
        result = search_primo(isbn_clean)
        isbn_results[isbn] = result

    # Map results back to the dataframe
    rows = []
    for _, row in df.iterrows():
        result = isbn_results.get(row['ISBN'])
        if result:
            in_catalog  = 'Yes'
            has_phys    = 'Yes' if result['Has_Physical']    else 'No'
            has_elec    = 'Yes' if result['Has_Electronic']  else 'No'
            cat_title   = result['Catalog_Title']
            cat_format  = result['Catalog_Format']
            cat_year    = result['Pub_Year']
            if result['Has_Physical'] or result['Has_Electronic']:
                action = 'In Catalog - Add to Reading List'
            else:
                action = 'In Catalog - Check Availability'
        else:
            in_catalog  = 'No'
            has_phys    = 'No'
            has_elec    = 'No'
            cat_title   = ''
            cat_format  = ''
            cat_year    = ''
            action      = 'Not in Catalog - Purchase Needed'
        rows.append({**row.to_dict(),
                     'In_Catalog':       in_catalog,
                     'Catalog_Title':    cat_title,
                     'Has_Physical':     has_phys,
                     'Has_Electronic':   has_elec,
                     'Catalog_Format':   cat_format,
                     'Catalog_Pub_Year': cat_year,
                     'Catalog_Action':   action})

    return pd.DataFrame(rows)


# ============================================================================
# PREPARE MERGED DATASET
# ============================================================================

def prepare_merged(merged):
    """
    Add helper columns to the merged Spring 2026 dataset.
    Also applies title- and publisher-based filters:
      - Excludes titles the library does not purchase (lab manuals, bundled access, etc.)
      - Excludes items from OER publishers (OpenStax, OpenTextbook, etc.)
      - Reclassifies OER/OpenStax titles from Book to 'OER / Open Educational Resource'
    """
    df = merged.copy()
    title_lower = df['Title'].str.lower().fillna('')

    # --- Exclude titles the library doesn't purchase ---
    exclude_title_mask = title_lower.apply(
        lambda t: any(term in t for term in EXCLUDE_TITLE_TERMS)
    )
    excluded_titles = exclude_title_mask.sum()
    if excluded_titles:
        print(f"  Excluded {excluded_titles} rows with non-purchasable title terms (lab manuals, bundled access, etc.)")
    df = df[~exclude_title_mask].copy()

    # --- Exclude items from OER publishers ---
    publisher_lower = df['Publisher'].str.lower().fillna('')
    exclude_pub_mask = publisher_lower.apply(
        lambda p: any(term in p for term in EXCLUDE_PUBLISHER_TERMS)
    )
    excluded_pubs = exclude_pub_mask.sum()
    if excluded_pubs:
        print(f"  Excluded {excluded_pubs} rows from OER publishers (OpenStax, OpenTextbook, etc.)")
    df = df[~exclude_pub_mask].copy()

    # Recompute title_lower after exclusions
    title_lower = df['Title'].str.lower().fillna('')

    # --- Reclassify OER/OpenStax as non-print ---
    oer_mask = (df['TextbookType'] == 'Book') & title_lower.apply(
        lambda t: any(term in t for term in OER_TITLE_TERMS)
    )
    reclassified = oer_mask.sum()
    if reclassified:
        print(f"  Reclassified {reclassified} OER/OpenStax titles from Book to non-print")
    df.loc[oer_mask, 'TextbookType'] = 'OER / Open Educational Resource'

    df['Course_Clean'] = df['Course'].str.strip()
    df['ISBN_Set']     = df['ISBN'].apply(extract_isbns)
    df['Title_Norm']   = df['Title'].apply(normalize_title)
    return df


# ============================================================================
# MATCH LOGIC
# ============================================================================

def match_merged_to_citations(merged_df, sp26_citations_df):
    """
    For each row in the merged dataset, try to find a matching citation
    in the Spring 2026 citations.

    Match priority:
      1. Course Code + ISBN overlap
      2. Course Code + normalized title
      3. ISBN overlap only (course code mismatch)

    Returns a list of result dicts.
    """
    # Build lookup structures from SP26 citations
    # Key: course_code -> list of citation rows
    cit_by_course = {}
    for _, row in sp26_citations_df.iterrows():
        course = row['Course_Clean']
        cit_by_course.setdefault(course, []).append(row)

    # Build ISBN -> citation lookup (for cross-course matching)
    cit_by_isbn = {}
    for _, row in sp26_citations_df.iterrows():
        for isbn in row['ISBN_Set']:
            cit_by_isbn.setdefault(isbn, []).append(row)

    results = []

    for _, mrow in merged_df.iterrows():
        course    = mrow['Course_Clean']
        isbn_set  = mrow['ISBN_Set']
        title_n   = mrow['Title_Norm']
        edition   = str(mrow.get('Edition', '')).strip()

        match_found   = False
        match_status  = ''
        match_notes   = ''
        matched_cit_title = ''
        matched_cit_course = ''

        # --- Attempt 1: Course + ISBN ---
        if course in cit_by_course and isbn_set:
            for cit in cit_by_course[course]:
                if isbns_overlap(isbn_set, cit['ISBN_Set']):
                    match_found  = True
                    match_status = 'Confirmed - ISBN Match'
                    matched_cit_title  = cit['Citation Title']
                    matched_cit_course = cit['Course Code']
                    break

        # --- Attempt 2: Course + normalized title ---
        if not match_found and course in cit_by_course and title_n:
            for cit in cit_by_course[course]:
                if title_n and cit['Title_Norm'] and title_n in cit['Title_Norm']:
                    match_found  = True
                    match_status = 'Confirmed - Title Match'
                    match_notes  = 'Verify ISBN'
                    matched_cit_title  = cit['Citation Title']
                    matched_cit_course = cit['Course Code']
                    break

        # --- Attempt 3: ISBN match under different course code ---
        if not match_found and isbn_set:
            for isbn in isbn_set:
                if isbn in cit_by_isbn:
                    for cit in cit_by_isbn[isbn]:
                        match_found  = True
                        match_status = 'ISBN Found - Different Course Code'
                        match_notes  = (
                            f"Merged course: {course} | "
                            f"Citation course: {cit['Course Code']}"
                        )
                        matched_cit_title  = cit['Citation Title']
                        matched_cit_course = cit['Course Code']
                        break
                if match_found:
                    break

        if not match_found:
            match_status = 'Missing from Reserves'

        results.append({
            'Course':              mrow.get('Course', ''),
            'Section':             mrow.get('Section', ''),
            'Instructor':          mrow.get('Instructor_Name', ''),
            'Total_Enrollment':    mrow.get('Total_Enrollment', ''),
            'Title':               mrow.get('Title', ''),
            'ISBN':                mrow.get('ISBN', ''),
            'Edition':             edition,
            'TextbookType':        mrow.get('TextbookType', ''),
            'Publisher':           mrow.get('Publisher', ''),
            'Match_Status':        match_status,
            'Match_Notes':         match_notes,
            'Matched_Citation_Title':  matched_cit_title,
            'Matched_Citation_Course': matched_cit_course,
        })

    return pd.DataFrame(results)


# ============================================================================
# REMOVAL CANDIDATES
# ============================================================================

def build_removal_list(citations_df):
    """
    Identify citations NOT in Spring 2026 that are candidates for removal.
    Only includes items whose Last_Active_Term is Spring 2024 or older
    (approx. 2 years before the current semester, Spring 2026).
    Sorted by: physical RESE shelf items first, then most recent term.
    """
    not_sp26 = citations_df[~citations_df['Is_SP26']].copy()

    # Assign a chronological rank to each item's most recent term
    not_sp26['Term_Rank'] = not_sp26['Most_Recent_Term'].map(TERM_ORDER).fillna(0)

    # Only include items last used at or before the cutoff (Spring 2024)
    not_sp26 = not_sp26[not_sp26['Term_Rank'] <= REMOVAL_CUTOFF_RANK]
    not_sp26 = not_sp26.sort_values(
        ['On_RESE_Shelf', 'Term_Rank'],
        ascending=[False, False]
    )

    removal = not_sp26[[
        'Citation Title', 'Course Code', 'Course_Clean', 'Course Terms',
        'Most_Recent_Term', 'Course Name', 'Academic Department',
        'Citation Status', 'On_RESE_Shelf', 'In_CLOSED_Location', 'Availability',
        'ISBN', 'ISBN (13)', 'MMS ID', 'Citation ID',
        'Reading List',
    ]].copy()

    removal = removal.rename(columns={
        'Citation Title':     'Title',
        'Course Code':        'Citation_Course_Code',
        'Course_Clean':       'Course_Code_Clean',
        'Most_Recent_Term':   'Last_Active_Term',
        'On_RESE_Shelf':      'On_RESE_Shelf',
        'In_CLOSED_Location': 'In_CLOSED_Location',
        'Availability':       'Shelf_Location',
    })

    # Add cleanup type label so staff know what action is required
    def assign_cleanup_type(row):
        if row['On_RESE_Shelf']:
            return 'PHYSICAL - Retrieve book from RESE shelf and reshelve to permanent location'
        elif row['In_CLOSED_Location']:
            return 'DIGITAL - Remove citation from reading list only (book stays in permanent CLOSED location)'
        else:
            return 'DIGITAL - Remove citation from reading list in Alma'

    removal['Cleanup_Type'] = removal.apply(assign_cleanup_type, axis=1)

    return removal


# ============================================================================
# EXCEL STYLING
# ============================================================================

# Color scheme per tab - (header_hex, zebra_row_hex)
TAB_COLORS = {
    'Summary':                 ('1F4E79', 'DCE6F1'),  # Dark blue / light blue
    'Removal_Priority_RESE':   ('C00000', 'FCE4D6'),  # Dark red / light peach
    'Removal_ReadingList_Only':('E26B0A', 'FFF2CC'),  # Dark orange / light yellow
    'Missing_Books':           ('833C00', 'FCE4D6'),  # Dark brown-red / light peach
    'Missing_NonPrint':        ('17375E', 'DDEBF7'),  # Dark teal / light blue
    'Confirmed_On_Reserves':   ('375623', 'E2EFDA'),  # Dark green / light green
    'Needs_Review':            ('7F6000', 'FFEB9C'),  # Dark gold / light yellow
    'Alma_No_Textbook_Request':  ('7030A0', 'EAD1FF'),  # Purple / light lavender
    'Roster_No_Alma_Course':     ('C00000', 'FCE4D6'),  # Dark red / light peach
    'Section_Instructor_Check':  ('7F6000', 'FFEB9C'),  # Dark gold / light yellow
}

# Colors for the Cleanup_Type column values
CLEANUP_FILLS = {
    'PHYSICAL': PatternFill('solid', fgColor='FCE4D6'),  # light red
    'DIGITAL':  PatternFill('solid', fgColor='DDEBF7'),  # light blue
}


def style_sheet(ws, tab_name):
    """
    Apply color formatting to a worksheet:
      - Colored header row with white bold text
      - Zebra-stripe alternating row shading
      - Freeze top row
      - Auto-fit column widths
      - For removal tabs: color the Cleanup_Type column by value
      - For Summary tab: bold the section header rows (those starting with '---')
    """
    header_hex, zebra_hex = TAB_COLORS.get(tab_name, ('1F4E79', 'DCE6F1'))
    header_fill = PatternFill('solid', fgColor=header_hex)
    zebra_fill  = PatternFill('solid', fgColor=zebra_hex)
    white_fill  = PatternFill('solid', fgColor='FFFFFF')
    header_font = Font(bold=True, color='FFFFFF')
    section_font = Font(bold=True, color=header_hex)

    # Find the Cleanup_Type column index if present
    cleanup_col_idx = None
    header_row = [cell.value for cell in ws[1]]
    if 'Cleanup_Type' in header_row:
        cleanup_col_idx = header_row.index('Cleanup_Type') + 1  # 1-based

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical='top')

        if row_idx == 1:
            # Header row
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
        else:
            # Check if this is a Summary section header (starts with '---')
            first_val = str(row[0].value or '')
            if tab_name == 'Summary' and first_val.startswith('---'):
                for cell in row:
                    cell.font = section_font
                    cell.fill = zebra_fill
            else:
                # Zebra striping
                fill = zebra_fill if row_idx % 2 == 0 else white_fill
                for cell in row:
                    cell.fill = fill

            # Override fill for Cleanup_Type column based on value
            if cleanup_col_idx:
                ctype_cell = row[cleanup_col_idx - 1]
                val = str(ctype_cell.value or '')
                if val.startswith('PHYSICAL'):
                    for cell in row:
                        cell.fill = CLEANUP_FILLS['PHYSICAL']
                elif val.startswith('DIGITAL'):
                    for cell in row:
                        cell.fill = CLEANUP_FILLS['DIGITAL']

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Auto-fit column widths (capped at 60)
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


# ============================================================================
# CONSOLIDATION HELPERS
# ============================================================================

def consolidate_missing(df, group_keys):
    """
    Collapse multiple section rows into one row per unique group_keys combination.
    - Sections are joined as comma-separated list
    - Instructors are joined as unique sorted list
    - Total_Enrollment is summed
    - Other fields (Edition, Publisher, TextbookType, etc.) taken from first row
    """
    def join_unique(series):
        vals = [str(v) for v in series.dropna().unique() if str(v).strip()]
        return ', '.join(sorted(vals))

    agg = df.groupby(group_keys, dropna=False).agg(
        Sections       =('Section',          join_unique),
        Instructors    =('Instructor',        join_unique),
        Total_Enrollment=('Total_Enrollment', 'sum'),
        Edition        =('Edition',           'first'),
        TextbookType   =('TextbookType',      'first'),
        Publisher      =('Publisher',         'first'),
        Match_Status   =('Match_Status',      'first'),
        Match_Notes    =('Match_Notes',       'first'),
    ).reset_index()

    # Reorder columns for readability
    front = list(group_keys) + ['Sections', 'Instructors', 'Total_Enrollment']
    rest  = [c for c in agg.columns if c not in front]
    return agg[front + rest]


# ============================================================================
# MAIN
# ============================================================================

def main():
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(REPORTS_DIR, f"reserves_comparison_{timestamp}.xlsx")

    print("=" * 65)
    print("COURSE RESERVES COMPARISON REPORT - CUNY BMCC Spring 2026")
    print("=" * 65)
    print()

    # Load
    citations, merged = load_data()

    # Course roster check via Alma Courses API
    print("Running course roster check against Alma...")
    api_key = load_alma_api_key()
    if api_key:
        alma_sp26_courses = fetch_active_sp26_courses(api_key)
    else:
        print("  WARNING: No Alma API key found in config.json. Skipping roster check.")
        alma_sp26_courses = []
    print()

    # Prepare
    print("Preparing data...")
    citations = prepare_citations(citations)
    merged    = prepare_merged(merged)

    sp26_cit = citations[citations['Is_SP26']].copy()
    print(f"  Spring 2026 citations: {len(sp26_cit)}")
    print(f"  Non-Spring-2026 citations: {len(citations) - len(sp26_cit)}")
    print()

    # Run matching
    print("Matching merged dataset against citations...")
    match_results = match_merged_to_citations(merged, sp26_cit)

    confirmed      = match_results[match_results['Match_Status'].str.startswith('Confirmed')]
    isbn_diff_crs  = match_results[match_results['Match_Status'] == 'ISBN Found - Different Course Code']
    missing        = match_results[match_results['Match_Status'] == 'Missing from Reserves']

    print(f"  Confirmed matches:                 {len(confirmed)}")
    print(f"  ISBN found, different course code: {len(isbn_diff_crs)}")
    print(f"  Missing from reserves:             {len(missing)}")
    print()

    # Course roster cross-reference
    if alma_sp26_courses:
        print("Cross-referencing course roster against textbook spreadsheet...")
        alma_no_request, roster_no_alma = build_course_roster_check(merged, alma_sp26_courses)
        print(f"  Active Alma SP26 courses with no textbook request: {len(alma_no_request):,}")
        print(f"  Textbook requests with no active Alma course:      {len(roster_no_alma):,}")
        print()
        print("Checking sections and instructors for matched courses...")
        section_instr_check = build_section_instructor_check(merged, alma_sp26_courses)
        print(f"  Matched courses with section/instructor mismatches: {len(section_instr_check):,}")
        print()
    else:
        alma_no_request     = pd.DataFrame()
        roster_no_alma      = pd.DataFrame()
        section_instr_check = pd.DataFrame()

    # Build removal list
    print("Building removal candidates list...")
    removal = build_removal_list(citations)
    removal_rese   = removal[removal['On_RESE_Shelf'] == True]
    removal_norese = removal[removal['On_RESE_Shelf'] == False]
    removal_closed = removal[removal['In_CLOSED_Location'] == True]
    print(f"  Not in Spring 2026 (total):        {len(removal)}")
    print(f"    - On temporary RESE shelf:       {len(removal_rese)} (physical retrieval needed)")
    print(f"    - In permanent CLOSED location:  {len(removal_closed)} (citation removal only)")
    print(f"    - Not on any reserve shelf:      {len(removal_norese) - len(removal_closed)} (citation removal only)")
    print()

    # -------------------------------------------------------------------------
    # Build Summary tab
    # -------------------------------------------------------------------------
    summary_data = {
        'Category': [
            '--- SPRING 2026 CITATIONS (CURRENT ALMA RESERVES) ---',
            'Total citations in Alma file',
            'Active for Spring 2026',
            'NOT in Spring 2026 (candidates for removal)',
            '  - On temporary RESE shelf (priority)',
            '  - Citation removal only (CLOSED permanent + non-shelf items)',
            '',
            '--- SPRING 2026 MERGED DATASET (WHAT SHOULD BE ON RESERVE) ---',
            'Total rows in merged dataset',
            'Unique courses needing reserves',
            '',
            '--- COMPARISON RESULTS ---',
            'Confirmed on reserves (ISBN match)',
            'Confirmed on reserves (title match)',
            'ISBN found but under different course code',
            'Missing from reserves - Books (see Missing_Books tab)',
            'Missing from reserves - Non-Print (see Missing_NonPrint tab)',
            '',
            '--- COURSES ---',
            'Unique courses with reserves set up (SP26 Citations)',
            'Unique courses in Spring 2026 merged dataset',
            'Courses in merged dataset with no reserves found',
            '',
            '--- COURSE ROSTER CHECK (ALMA COURSES API) ---',
            'Active Spring 2026 courses in Alma',
            'Active Alma courses with NO textbook request (see Alma_No_Textbook_Request)',
            'Textbook requests with NO active Alma course (see Roster_No_Alma_Course)',
            'Matched courses with section or instructor mismatches (see Section_Instructor_Check)',
        ],
        'Count': [
            '',
            len(citations),
            len(sp26_cit),
            len(removal),
            len(removal_rese),
            len(removal_norese),
            '',
            '',
            len(merged),
            merged['Course_Clean'].nunique(),
            '',
            '',
            len(confirmed[confirmed['Match_Status'] == 'Confirmed - ISBN Match']),
            len(confirmed[confirmed['Match_Status'] == 'Confirmed - Title Match']),
            len(isbn_diff_crs),
            len(missing[missing['TextbookType'] == 'Book']),
            len(missing[missing['TextbookType'] != 'Book']),
            '',
            '',
            sp26_cit['Course_Clean'].nunique(),
            merged['Course_Clean'].nunique(),
            missing['Course'].nunique() if len(missing) > 0 else 0,
            '',
            '',
            len(alma_sp26_courses),
            len(alma_no_request),
            len(roster_no_alma),
            len(section_instr_check),
        ],
        'Notes': [
            '',
            'All citations exported from Alma',
            'Course Terms field includes "Spring 2026"',
            'Course Terms does NOT include "Spring 2026"',
            'PHYSICAL - retrieve from RESE shelf and reshelve to permanent location',
            'DIGITAL only - includes CLOSED (permanent) and non-shelf items',
            '',
            '',
            'All Spring 2026 course/textbook requests',
            'Distinct course codes',
            '',
            '',
            'Matched by ISBN - high confidence',
            'Matched by title - verify ISBN',
            'Same book, different course code - review needed',
            'TextbookType = Book - physical reserves action needed',
            'E-Books, E-Resources, Recordings, etc.',
            '',
            '',
            'Has at least 1 citation active for Spring 2026',
            'Has at least 1 textbook request for Spring 2026',
            '',
            '',
            '',
            'Fetched live from Alma Courses API',
            'Faculty may not have submitted textbook info yet',
            'Course may need to be created or reactivated in Alma',
            'Section or instructor data differs between spreadsheet and Alma',
        ]
    }
    df_summary = pd.DataFrame(summary_data)

    # -------------------------------------------------------------------------
    # Write Excel
    # -------------------------------------------------------------------------
    print(f"Writing report to: {output_file}")
    os.makedirs(REPORTS_DIR, exist_ok=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        # Tab 1: Summary
        df_summary.to_excel(writer, sheet_name='Summary', index=False)

        # Tab 2: Removal - Priority (physically on RESE shelf)
        fix_id_columns(removal_rese).to_excel(
            writer, sheet_name='Removal_Priority_RESE', index=False
        )

        # Tab 3: Removal - Reading list only (not on physical shelf)
        fix_id_columns(removal_norese).to_excel(
            writer, sheet_name='Removal_ReadingList_Only', index=False
        )

        # Tab 4a: Missing from reserves - Books, consolidated by ISBN + Course
        missing_cols = [
            'Course', 'Section', 'Instructor', 'Total_Enrollment',
            'Title', 'ISBN', 'Edition', 'TextbookType', 'Publisher',
            'Match_Status', 'Match_Notes',
        ]
        missing_books_raw = missing[missing['TextbookType'] == 'Book'][missing_cols]
        missing_books = consolidate_missing(missing_books_raw, ['Course', 'ISBN', 'Title'])
        missing_books = lookup_missing_books_in_primo(missing_books)
        missing_books.to_excel(writer, sheet_name='Missing_Books', index=False)

        # Tab 4b: Missing from reserves - Non-print, consolidated by Title + Course
        missing_nonprint_raw = missing[missing['TextbookType'] != 'Book'][missing_cols]
        missing_nonprint = consolidate_missing(missing_nonprint_raw, ['Course', 'Title'])
        missing_nonprint.to_excel(writer, sheet_name='Missing_NonPrint', index=False)

        # Tab 5: Confirmed correct
        confirmed_out = confirmed[[
            'Course', 'Section', 'Instructor', 'Total_Enrollment',
            'Title', 'ISBN', 'Edition', 'TextbookType',
            'Match_Status', 'Matched_Citation_Title', 'Matched_Citation_Course',
            'Match_Notes',
        ]]
        confirmed_out.to_excel(writer, sheet_name='Confirmed_On_Reserves', index=False)

        # Tab 6: Needs review (ISBN found, different course code)
        if len(isbn_diff_crs) > 0:
            review_out = isbn_diff_crs[[
                'Course', 'Section', 'Instructor', 'Total_Enrollment',
                'Title', 'ISBN', 'Edition', 'TextbookType',
                'Match_Status', 'Match_Notes',
                'Matched_Citation_Title', 'Matched_Citation_Course',
            ]]
            review_out.to_excel(writer, sheet_name='Needs_Review', index=False)

        # Tab 7: Active Alma courses with no textbook request
        if not alma_no_request.empty:
            alma_no_request.to_excel(
                writer, sheet_name='Alma_No_Textbook_Request', index=False
            )

        # Tab 8: Textbook requests with no active Alma course
        if not roster_no_alma.empty:
            roster_no_alma.to_excel(
                writer, sheet_name='Roster_No_Alma_Course', index=False
            )

        # Tab 9: Section and instructor mismatches for matched courses
        if not section_instr_check.empty:
            section_instr_check.to_excel(
                writer, sheet_name='Section_Instructor_Check', index=False
            )

        # Apply color formatting to all tabs
        for tab_name, ws in writer.sheets.items():
            style_sheet(ws, tab_name)

    print()
    print("=" * 65)
    print("DONE")
    print("=" * 65)
    print(f"Report saved to: {output_file}")
    print()
    print("Tabs in the report:")
    print("  Summary                   - Overview counts and categories")
    print("  Removal_Priority_RESE     - Items physically on reserve shelf, NOT Spring 2026")
    print("  Removal_ReadingList_Only  - Items in reading lists only, NOT Spring 2026")
    print("  Missing_Books             - Spring 2026 Books not found in Alma citations")
    print("  Missing_NonPrint          - Spring 2026 Non-Print items not found in Alma citations")
    print("  Confirmed_On_Reserves     - Items correctly matched")
    print("  Needs_Review              - ISBN found but course code differs")
    print("  Alma_No_Textbook_Request  - Active Alma courses with no textbook request submitted")
    print("  Roster_No_Alma_Course     - Textbook requests with no matching active Alma course")
    print("  Section_Instructor_Check  - Matched courses where sections or instructors differ")
    print()
    print("NO CHANGES have been made to Alma.")


if __name__ == "__main__":
    main()
