"""
Old Textbooks Collection Cleanup - CUNY BMCC
---------------------------------------------
Applies library policy to identify items to weed from the Old Textbooks
collection and generates a report. NO changes are made to Alma.

INPUT FILE
----------
CSV exported from Alma Analytics using the Physical Items subject area.
Filter logic (OR conditions):
  - Permanent Location Code = OLDTX
  - OR Permanent Location Code = CLOSED
  - OR Temporary Location Code = RESE

Columns required:
  MMS Id, Lifecycle, Location Code, Temporary Location Code, Title,
  Title (Normalized), Publication Date, Due Back Date (calendar),
  Barcode, Physical Item Id, Base Status, Process Type, Copy ID,
  Holdings ID, Permanent Call Number, Normalized Call Number

LOCATION POLICY
---------------
  OLDTX (Old Textbooks permanent location):
    - Included in edition ranking
    - Included in copy count (MAX_COPIES=3)
    - Items with old editions → Weed_Old_Edition tab
    - Excess copies → Weed_Excess_Copy tab
    - Missing/Lost items → Weed_Missing_Lost tab

  All other locations (STACK, CLOSED, COMIC, DICT, BMPB, etc.):
    - Included in edition ranking (their copies count toward "is this edition active?")
    - NOT included in copy count
    - Items with old editions → Needs_Review tab (no weed recommendation)

MISSING/LOST DETECTION
-----------------------
Uses the Process Type column (NOT Base Status alone):
  - Missing/Lost = Process Type in {Lost, Missing, Lost and paid}
  - On-loan = Process Type in {Loan, Transit, In Process, Acquisition}
    → On-loan items COUNT as real copies (they will come back)
  - Missing/Lost items are excluded from edition ranking and copy count

EDITION RANKING
----------------
Editions are grouped by call number STEM (year stripped from call number).
e.g. "RT41 .K74 2019" → stem "RT41 .K74"
Within each stem group, editions are ranked newest-first.
Only editions with at least one non-missing/lost copy (in any location)
occupy an edition slot.

HEALTH BOOK CLASSIFICATION
---------------------------
Health/biology books keep 2 editions; non-health keep 3.
Health = LCC R class (medicine/nursing) + QH/QK/QL/QM/QP/QR (biology).

Usage:
    python scripts/old_textbooks_cleanup.py

Author: Patty (with Claude Code assistance)
Date: 2026-03-05
"""

import pandas as pd
import re
from datetime import datetime
import os
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ============================================================================
# FILE PATHS
# ============================================================================

BASE_DIR    = "/Users/patty_home/Desktop/Agentic AI/Reserves Tool"
INPUT_FILE  = os.path.join(BASE_DIR, "data/cleanup4_reserves_textbooks.csv")
REPORTS_DIR = os.path.join(BASE_DIR, "reports")

# ============================================================================
# POLICY CONSTANTS
# ============================================================================

MAX_COPIES              = 3   # Maximum copies kept per edition (OLDTX only)
MAX_EDITIONS_NON_HEALTH = 3   # Keep latest N editions for non-health titles
MAX_EDITIONS_HEALTH     = 2   # Keep latest N editions for health/biology titles

# Location that receives weed recommendations
OLDTX_LOCATION = 'OLDTX'

# Process Type values that mean the item is truly missing/lost (not just on loan)
MISSING_LOST_PROCESS_TYPES = {'Lost', 'Missing', 'Lost and paid'}

# LCC call number prefixes that indicate health/biology subjects
# Matches: R, RA, RB, RC, RD, RE, RF, RG, RJ, RK, RL, RM, RS, RT, RV, RX, RZ
#          QH, QK, QL, QM, QP, QR
HEALTH_LCC_PATTERN = re.compile(
    r'^(R[ABCDEFGHJKLMSTUVXZ]?\d|Q[HKLMPR]\d)',
    re.IGNORECASE
)

# ============================================================================
# HELPERS
# ============================================================================

def call_stem(call_number):
    """
    Strip the trailing publication year from a call number to get the
    'base' identifier for a work across editions.
    e.g. 'RT41 .K74 2019' -> 'rt41 .k74'
         'Z246 .M43 2016' -> 'z246 .m43'
    """
    if pd.isna(call_number) or str(call_number).strip() in ('', 'Unknown'):
        return ''
    s = re.sub(r'\s+(?:19|20)\d{2}.*$', '', str(call_number)).strip()
    return s.lower()


def extract_pub_year(call_number, pub_date=''):
    """
    Extract the publication year from a call number or Publication Date field.
    e.g. 'RT41 .K74 2019' -> 2019
         '[2021]'          -> 2021
         '©2004'           -> 2004
    Returns int year or None if not found.
    """
    years = re.findall(r'\b((?:19|20)\d{2})\b', str(call_number or ''))
    if years:
        return int(years[-1])
    years = re.findall(r'\b((?:19|20)\d{2})\b', str(pub_date or ''))
    if years:
        return int(years[-1])
    return None


def is_health_book(call_number):
    """
    Return True if the LCC call number indicates health or biology.
    Health = R class (medicine/nursing) or Q biology subclasses.
    """
    if pd.isna(call_number) or str(call_number).strip() in ('', 'Unknown'):
        return False
    return bool(HEALTH_LCC_PATTERN.match(str(call_number).strip()))


def is_missing_lost(process_type):
    """
    Return True if Process Type indicates truly missing/lost (not on loan).
    On-loan items (Loan, Transit, In Process, Acquisition) are NOT missing —
    they will return and should count as real copies.
    """
    return str(process_type).strip() in MISSING_LOST_PROCESS_TYPES


# ============================================================================
# CORE ANALYSIS
# ============================================================================

def analyze_collection(df):
    """
    Apply all policy rules to every item.

    Edition ranking uses ALL locations (OLDTX + STACK + CLOSED + etc.).
    Copy count and weed recommendations apply to OLDTX items only.
    Non-OLDTX items with old editions go to the Needs_Review tab.

    Returns item-level DataFrame with analysis columns added.
    """

    df = df.copy()

    # --- Step 0: Key flags per item ---
    df['Is_Missing_Lost'] = df['Process Type'].apply(is_missing_lost)
    df['Is_OLDTX']        = df['Location Code'] == OLDTX_LOCATION
    # An "active copy" counts toward edition ranking and copy limits
    df['Is_Active_Copy']  = ~df['Is_Missing_Lost']

    # --- Step 1: Build bib-level (MMS Id) summary across ALL locations ---
    mms = df.groupby('MMS Id').agg(
        Title             =('Title',                'first'),
        Call_Number       =('Permanent Call Number', 'first'),
        Pub_Date          =('Publication Date',      'first'),
        Total_Item_Count  =('Physical Item Id',      'count'),
        Active_Copy_Count =('Is_Active_Copy',        'sum'),   # non-missing/lost copies anywhere
    ).reset_index()

    mms['Call_Stem'] = mms['Call_Number'].apply(call_stem)
    mms['Pub_Year']  = mms.apply(
        lambda r: extract_pub_year(r['Call_Number'], r['Pub_Date']), axis=1
    )
    mms['Is_Health'] = mms['Call_Number'].apply(is_health_book)

    # --- Step 2: Within each call-stem group, rank editions newest-first ---
    # Only editions with at least one active (non-missing/lost) copy anywhere
    # count toward the edition limit.
    stem_year_rows = []
    for stem, group in mms.groupby('Call_Stem'):
        if stem == '':
            continue
        active_group = group[group['Active_Copy_Count'] > 0]
        unique_years = sorted(
            [y for y in active_group['Pub_Year'].dropna().unique()],
            reverse=True
        )
        total = len(unique_years)
        for rank, yr in enumerate(unique_years, start=1):
            stem_year_rows.append({
                'Call_Stem':      stem,
                'Pub_Year':       yr,
                'Edition_Rank':   rank,
                'Total_Editions': total,
            })
    stem_year_df = (
        pd.DataFrame(stem_year_rows) if stem_year_rows
        else pd.DataFrame(columns=['Call_Stem', 'Pub_Year', 'Edition_Rank', 'Total_Editions'])
    )

    mms = mms.merge(stem_year_df, on=['Call_Stem', 'Pub_Year'], how='left')

    # --- Step 3: Flag editions that exceed the keep limit ---
    def edition_exceeds_limit(row):
        if pd.isna(row.get('Edition_Rank')):
            return False
        max_ed = MAX_EDITIONS_HEALTH if row['Is_Health'] else MAX_EDITIONS_NON_HEALTH
        return int(row['Edition_Rank']) > max_ed

    mms['Edition_Weed'] = mms.apply(edition_exceeds_limit, axis=1)

    # --- Step 4: Merge bib-level info back to item level ---
    item_df = df.merge(
        mms[['MMS Id', 'Call_Stem', 'Pub_Year', 'Is_Health',
             'Active_Copy_Count', 'Edition_Rank', 'Total_Editions', 'Edition_Weed']],
        on='MMS Id',
        how='left'
    )

    # --- Step 5: Within each MMS Id, rank OLDTX active copies only ---
    # Copy count only applies to OLDTX; missing/lost copies excluded.
    item_df = item_df.sort_values(['MMS Id', 'Barcode']).reset_index(drop=True)
    oldtx_active_mask = item_df['Is_OLDTX'] & item_df['Is_Active_Copy']
    item_df['Copy_Rank'] = 0
    item_df.loc[oldtx_active_mask, 'Copy_Rank'] = (
        item_df.loc[oldtx_active_mask].groupby('MMS Id').cumcount() + 1
    )
    # Total_Copies = OLDTX active copies per MMS Id
    oldtx_active_count = (
        item_df[item_df['Is_OLDTX'] & item_df['Is_Active_Copy']]
        .groupby('MMS Id')['Is_Active_Copy']
        .sum()
        .rename('Total_Copies')
    )
    item_df = item_df.merge(oldtx_active_count, on='MMS Id', how='left')
    item_df['Total_Copies'] = item_df['Total_Copies'].fillna(0).astype(int)

    # --- Step 6: Flag excess OLDTX copies for editions that are NOT being weeded ---
    item_df['Copy_Weed'] = (
        item_df['Is_OLDTX'] &
        item_df['Is_Active_Copy'] &
        (~item_df['Edition_Weed']) &
        (item_df['Copy_Rank'] > MAX_COPIES)
    )

    # --- Step 7: Flag OLDTX missing/lost items (not already caught by Edition_Weed) ---
    item_df['Missing_Weed'] = (
        item_df['Is_OLDTX'] &
        item_df['Is_Missing_Lost'] &
        (~item_df['Edition_Weed'])
    )

    # --- Step 8: Flag non-OLDTX items with old editions for review ---
    item_df['Needs_Review'] = (
        (~item_df['Is_OLDTX']) &
        item_df['Edition_Weed']
    )

    # --- Step 9: Assign Action and Weed_Reason ---
    def weed_reason(row):
        if row.get('Edition_Weed') and row.get('Is_OLDTX'):
            health_str = 'Health/Biology' if row['Is_Health'] else 'Non-Health'
            max_ed     = MAX_EDITIONS_HEALTH if row['Is_Health'] else MAX_EDITIONS_NON_HEALTH
            yr         = int(row['Pub_Year'])       if pd.notna(row.get('Pub_Year'))       else '?'
            rank       = int(row['Edition_Rank'])   if pd.notna(row.get('Edition_Rank'))   else '?'
            total      = int(row['Total_Editions']) if pd.notna(row.get('Total_Editions')) else '?'
            return (f"Old Edition ({health_str}) — "
                    f"pub year {yr} is rank {rank} of {total} editions "
                    f"(keep latest {max_ed})")
        if row.get('Missing_Weed'):
            pt = row.get('Process Type', '')
            return f"Missing/Lost — Process Type: {pt}"
        if row.get('Copy_Weed'):
            return (f"Excess Copy — copy {int(row['Copy_Rank'])} of {int(row['Total_Copies'])} "
                    f"OLDTX active copies (keep max {MAX_COPIES})")
        if row.get('Needs_Review'):
            health_str = 'Health/Biology' if row['Is_Health'] else 'Non-Health'
            max_ed     = MAX_EDITIONS_HEALTH if row['Is_Health'] else MAX_EDITIONS_NON_HEALTH
            yr         = int(row['Pub_Year'])       if pd.notna(row.get('Pub_Year'))       else '?'
            rank       = int(row['Edition_Rank'])   if pd.notna(row.get('Edition_Rank'))   else '?'
            total      = int(row['Total_Editions']) if pd.notna(row.get('Total_Editions')) else '?'
            loc        = row.get('Location Code', '')
            return (f"Review — non-OLDTX ({loc}) old edition ({health_str}): "
                    f"pub year {yr} is rank {rank} of {total} (keep latest {max_ed})")
        return ''

    item_df['Weed_Reason']     = item_df.apply(weed_reason, axis=1)
    item_df['Is_Health_Label'] = item_df['Is_Health'].apply(
        lambda x: 'Yes — Health/Biology' if x else 'No'
    )

    # Action: OLDTX items get WEED/KEEP; non-OLDTX items get REVIEW or KEEP
    def action(row):
        if row.get('Edition_Weed') and row.get('Is_OLDTX'):
            return 'WEED'
        if row.get('Copy_Weed') or row.get('Missing_Weed'):
            return 'WEED'
        if row.get('Needs_Review'):
            return 'REVIEW'
        return 'KEEP'

    item_df['Action'] = item_df.apply(action, axis=1)

    return item_df


# ============================================================================
# OUTPUT COLUMNS
# ============================================================================

def build_report_df(item_df):
    """Select and order columns for the output Excel tabs."""
    orig_cols = [
        'Title', 'Permanent Call Number', 'Barcode', 'Base Status', 'Process Type',
        'Location Code', 'Temporary Location Code',
        'Publication Date', 'Copy ID', 'MMS Id', 'Holdings ID', 'Physical Item Id',
    ]
    analysis_cols = [
        'Action', 'Weed_Reason',
        'Is_Health_Label', 'Pub_Year', 'Call_Stem',
        'Edition_Rank', 'Total_Editions',
        'Copy_Rank', 'Total_Copies',
    ]
    available = [c for c in orig_cols if c in item_df.columns]
    out = item_df[analysis_cols + available].copy()
    out = out.rename(columns={
        'Is_Health_Label':       'Is_Health_Subject',
        'Permanent Call Number': 'Call_Number',
    })
    # These columns were already read as strings from CSV to prevent scientific notation
    # but apply str conversion defensively for any numeric leakage
    for col in ['Barcode', 'MMS Id', 'Holdings ID', 'Physical Item Id', 'Copy ID']:
        if col in out.columns:
            out[col] = out[col].apply(
                lambda x: str(x) if pd.notna(x) and str(x) not in ('', 'nan') else ''
            )
    return out


# ============================================================================
# EXCEL STYLING
# ============================================================================

TAB_COLORS = {
    'Summary':           ('1F4E79', 'DCE6F1'),
    'Weed_Old_Edition':  ('C00000', 'FCE4D6'),
    'Weed_Excess_Copy':  ('E26B0A', 'FFF2CC'),
    'Weed_Missing_Lost': ('595959', 'F2F2F2'),
    'All_Weed':          ('833C00', 'FCE4D6'),
    'Needs_Review':      ('7F6000', 'FFEB9C'),
    'Keep':              ('375623', 'E2EFDA'),
    'No_Year_Found':     ('7F6000', 'FFEB9C'),
}

ACTION_FILLS = {
    'WEED':   PatternFill('solid', fgColor='FCE4D6'),
    'REVIEW': PatternFill('solid', fgColor='FFEB9C'),
    'KEEP':   PatternFill('solid', fgColor='E2EFDA'),
}


def style_sheet(ws, tab_name):
    header_hex, zebra_hex = TAB_COLORS.get(tab_name, ('1F4E79', 'DCE6F1'))
    header_fill  = PatternFill('solid', fgColor=header_hex)
    zebra_fill   = PatternFill('solid', fgColor=zebra_hex)
    white_fill   = PatternFill('solid', fgColor='FFFFFF')
    header_font  = Font(bold=True, color='FFFFFF')
    section_font = Font(bold=True, color=header_hex)

    header_row = [cell.value for cell in ws[1]]
    action_col = (header_row.index('Action') + 1) if 'Action' in header_row else None

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical='top')

        if row_idx == 1:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
        else:
            first_val = str(row[0].value or '')
            if tab_name == 'Summary' and first_val.startswith('---'):
                for cell in row:
                    cell.font = section_font
                    cell.fill = zebra_fill
            else:
                fill = zebra_fill if row_idx % 2 == 0 else white_fill
                for cell in row:
                    cell.fill = fill
                if action_col:
                    av = str(row[action_col - 1].value or '')
                    if av in ACTION_FILLS:
                        for cell in row:
                            cell.fill = ACTION_FILLS[av]

    ws.freeze_panes = 'A2'
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or '')) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 70)


# ============================================================================
# MAIN
# ============================================================================

def main():
    timestamp   = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(REPORTS_DIR, f"old_textbooks_cleanup_{timestamp}.xlsx")

    print("=" * 65)
    print("OLD TEXTBOOKS COLLECTION CLEANUP - CUNY BMCC")
    print("=" * 65)
    print()
    print(f"Policy:")
    print(f"  Max copies per edition (OLDTX only):  {MAX_COPIES}")
    print(f"  Max editions kept (non-health):        {MAX_EDITIONS_NON_HEALTH}")
    print(f"  Max editions kept (health/bio):        {MAX_EDITIONS_HEALTH}")
    print()

    # Read CSV — force string dtype on large integer ID columns to prevent
    # scientific notation in Excel output
    dtype_overrides = {
        'MMS Id':          str,
        'Barcode':         str,
        'Physical Item Id': str,
        'Holdings ID':     str,
        'Copy ID':         str,
    }
    print(f"Loading: {INPUT_FILE}")
    df = pd.read_csv(INPUT_FILE, encoding='utf-8-sig', dtype=dtype_overrides)

    total_items = len(df)
    unique_mms  = df['MMS Id'].nunique()
    loc_counts  = df['Location Code'].value_counts().to_dict()
    print(f"  Total items:        {total_items}")
    print(f"  Unique MMS IDs:     {unique_mms}")
    print(f"  Location breakdown: {loc_counts}")
    print()

    print("Analyzing collection...")
    item_df = analyze_collection(df)

    # OLDTX items only for weed tabs
    oldtx_df     = item_df[item_df['Is_OLDTX']]
    weed_df      = oldtx_df[oldtx_df['Action'] == 'WEED']
    weed_ed      = oldtx_df[oldtx_df['Edition_Weed']  == True]
    weed_cp      = oldtx_df[oldtx_df['Copy_Weed']     == True]
    weed_missing = oldtx_df[oldtx_df['Missing_Weed']  == True]
    keep_df      = oldtx_df[oldtx_df['Action'] == 'KEEP']
    no_year_df   = keep_df[keep_df['Pub_Year'].isna()]

    # Non-OLDTX items flagged for review
    review_df    = item_df[item_df['Needs_Review'] == True]

    weed_ed_health    = weed_ed[weed_ed['Is_Health'] == True]
    weed_ed_nonhealth = weed_ed[weed_ed['Is_Health'] == False]

    print(f"  OLDTX items to WEED:             {len(weed_df)}")
    print(f"    - Old edition (health/bio):    {len(weed_ed_health)}")
    print(f"    - Old edition (non-health):    {len(weed_ed_nonhealth)}")
    print(f"    - Excess copies (active):      {len(weed_cp)}")
    print(f"    - Missing/lost items:          {len(weed_missing)}")
    print(f"  OLDTX items to KEEP:             {len(keep_df)}")
    print(f"    - No pub year found (kept):    {len(no_year_df)}")
    print(f"  Non-OLDTX items for review:      {len(review_df)}")
    print()

    # -------------------------------------------------------------------------
    # Summary tab
    # -------------------------------------------------------------------------
    summary_rows = [
        ('--- COLLECTION OVERVIEW ---',                   '',                      ''),
        ('Total physical items loaded',                   total_items,             'All locations from the Alma Analytics report'),
        ('Unique bibliographic records (MMS IDs)',        unique_mms,              'One MMS ID = one edition / bibliographic record'),
        ('OLDTX items',                                   loc_counts.get('OLDTX', 0), 'Old Textbooks permanent location — weed candidates'),
        ('Non-OLDTX items (STACK, CLOSED, etc.)',         total_items - loc_counts.get('OLDTX', 0), 'Included for edition ranking only; see Needs_Review tab'),
        ('',                                              '',                      ''),
        ('--- POLICY APPLIED ---',                        '',                      ''),
        ('Missing/lost detection',                        'Process Type column',   'Lost, Missing, Lost and paid → excluded from counts'),
        ('On-loan items (Loan, Transit, etc.)',            'Count as active copies','These items will return; they hold a copy slot'),
        ('Max active copies kept per edition (OLDTX)',    MAX_COPIES,              'Only OLDTX non-missing copies count toward this limit'),
        ('Max editions kept — non-health books',          MAX_EDITIONS_NON_HEALTH, 'Non-health = all LCC classes except R and Q biology subclasses'),
        ('Max editions kept — health / biology books',    MAX_EDITIONS_HEALTH,     'Health = R class (medicine/nursing) + QH/QK/QL/QM/QP/QR (biology)'),
        ('Edition grouping method',                       'Call number stem',      'Year stripped from call number; e.g. "RT41 .K74 2019" → stem "RT41 .K74"'),
        ('Edition ranking scope',                         'All locations',         'Non-missing copies from OLDTX, STACK, CLOSED etc. all count'),
        ('',                                              '',                      ''),
        ('--- OLDTX WEED CANDIDATES ---',                 '',                      ''),
        ('Total OLDTX items flagged for WEEDING',         len(weed_df),            'NO changes made to Alma — this is a review list only'),
        ('  Old Edition — Health/Biology',                len(weed_ed_health),     f'Exceed {MAX_EDITIONS_HEALTH}-edition limit; withdraw all copies of those editions'),
        ('  Old Edition — Non-Health',                    len(weed_ed_nonhealth),  f'Exceed {MAX_EDITIONS_NON_HEALTH}-edition limit; withdraw all copies of those editions'),
        ('  Excess Active Copies (of kept editions)',      len(weed_cp),            f'Copy rank {MAX_COPIES + 1}+ among OLDTX active copies'),
        ('  Missing/Lost items',                          len(weed_missing),       'Process Type is Lost, Missing, or Lost and paid'),
        ('',                                              '',                      ''),
        ('--- OLDTX ITEMS TO KEEP ---',                   '',                      ''),
        ('Total OLDTX items to KEEP',                     len(keep_df),            'Items that comply with current policy'),
        ('  Items kept but no year found (review these)', len(no_year_df),         'Manually verify edition and health status'),
        ('',                                              '',                      ''),
        ('--- NON-OLDTX ITEMS ---',                       '',                      ''),
        ('Non-OLDTX items with old editions (review)',    len(review_df),          'STACK, CLOSED, etc. items whose edition exceeds the limit — no weed action, but review'),
        ('',                                              '',                      ''),
        ('--- HOW TO USE THIS REPORT ---',                '',                      ''),
        ('Weed_Old_Edition tab',                          len(weed_ed),            'OLDTX items from editions beyond the limit — withdraw all copies'),
        ('Weed_Excess_Copy tab',                          len(weed_cp),            'OLDTX excess active copies of kept editions — withdraw only flagged copies'),
        ('Weed_Missing_Lost tab',                         len(weed_missing),       'OLDTX missing/lost items — update status or withdraw in Alma'),
        ('All_Weed tab',                                  len(weed_df),            'Combined list of all OLDTX items to weed'),
        ('Needs_Review tab',                              len(review_df),          'Non-OLDTX items with old editions — no automatic weed, review with staff'),
        ('Keep tab',                                      len(keep_df),            'OLDTX items that meet policy — no action needed'),
        ('No_Year_Found tab',                             len(no_year_df),         'OLDTX items kept but pub year not detected — review manually'),
    ]
    df_summary = pd.DataFrame(summary_rows, columns=['Category', 'Count', 'Notes'])

    # -------------------------------------------------------------------------
    # Build report dataframes
    # -------------------------------------------------------------------------
    rep_ed      = build_report_df(weed_ed).sort_values(['Call_Stem', 'Pub_Year'])
    rep_cp      = build_report_df(weed_cp).sort_values(['Call_Stem', 'Copy_Rank'])
    rep_missing = build_report_df(weed_missing).sort_values(['Call_Stem', 'Pub_Year'])
    rep_weed    = build_report_df(weed_df).sort_values(['Weed_Reason', 'Call_Stem', 'Pub_Year'])
    rep_review  = build_report_df(review_df).sort_values(['Location Code', 'Call_Stem', 'Pub_Year'])
    rep_keep    = build_report_df(keep_df).sort_values(['Call_Stem', 'Pub_Year'])
    rep_noyr    = build_report_df(no_year_df).sort_values(['Title'])

    # -------------------------------------------------------------------------
    # Write Excel
    # -------------------------------------------------------------------------
    print(f"Writing report to: {output_file}")
    os.makedirs(REPORTS_DIR, exist_ok=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='Summary',           index=False)
        rep_ed.to_excel(    writer, sheet_name='Weed_Old_Edition',  index=False)
        rep_cp.to_excel(    writer, sheet_name='Weed_Excess_Copy',  index=False)
        rep_missing.to_excel(writer, sheet_name='Weed_Missing_Lost', index=False)
        rep_weed.to_excel(  writer, sheet_name='All_Weed',          index=False)
        rep_review.to_excel(writer, sheet_name='Needs_Review',      index=False)
        rep_keep.to_excel(  writer, sheet_name='Keep',              index=False)
        if len(rep_noyr) > 0:
            rep_noyr.to_excel(writer, sheet_name='No_Year_Found',   index=False)

        for tab_name, ws in writer.sheets.items():
            style_sheet(ws, tab_name)

    print()
    print("=" * 65)
    print("DONE")
    print("=" * 65)
    print(f"Report saved to: {output_file}")
    print()
    print("Tabs in the report:")
    print(f"  Summary            - Policy overview and counts")
    print(f"  Weed_Old_Edition   - {len(weed_ed)} OLDTX items from editions beyond the keep limit")
    print(f"  Weed_Excess_Copy   - {len(weed_cp)} excess OLDTX copies of kept editions")
    print(f"  Weed_Missing_Lost  - {len(weed_missing)} OLDTX items that are missing/lost")
    print(f"  All_Weed           - {len(weed_df)} total OLDTX items flagged for weeding")
    print(f"  Needs_Review       - {len(review_df)} non-OLDTX items with old editions (no weed action)")
    print(f"  Keep               - {len(keep_df)} OLDTX items that comply with policy")
    if len(rep_noyr) > 0:
        print(f"  No_Year_Found      - {len(no_year_df)} OLDTX items with no year detected (review manually)")
    print()
    print("NO CHANGES have been made to Alma.")


if __name__ == "__main__":
    main()
